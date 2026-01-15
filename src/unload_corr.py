# -*- coding: utf-8 -*-
"""
Модуль логики блока 2: Создание отчета по корректировке CF16
Аналог процедуры UnloadCorr из VBA
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from src.excel_handler import ExcelHandler, SheetHelper, TableFormatter
from src.helpers import search_arr_col, format_table, safe_str, safe_float
from src.logger import log, timing, Timer


# Названия листов
SHEET_NAMES = {
    "start": "СТАРТ",
    "marja": "МАРЖА",
    "correction_template": "КОРРЕКТИРОВКА_ШАБЛОН",
    "correction": "Корректировка_CF16",
}

# Названия расчётных листов
CALC_SHEETS = [
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_КЦ",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_ЦОД",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_КЦ",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_ЦОД",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_T2_КЦ",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_T2_ЦОД",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_T2_КЦ",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_T2_ЦОД",
]


@dataclass
class ProcessingResult:
    """Результат обработки"""

    success: bool = False
    message: str = ""
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


@dataclass
class CorrectionColumnInfo:
    """Информация о колонках листа Корректировка"""

    balance_unit: int = 0  # Балансовая Единица
    article: int = 0  # Статья
    cfo: int = 0  # ЦФО
    debit: int = 0  # Дебет
    interco2: int = 0  # Интерко2
    corrections: int = 0  # Корректировки
    account: int = 0  # Счёт
    description: int = 0  # Описание строки


class UnloadCorrProcessor:
    """
    Процессор для создания отчёта по корректировке CF16
    """

    def __init__(self, macros_file_path: str, progress_callback=None):
        """
        Args:
            macros_file_path: Путь к файлу с макросами (основной файл)
            progress_callback: Функция обратного вызова для обновления прогресса
        """
        self.macros_file_path = Path(macros_file_path)
        self.progress_callback = progress_callback

        self.wb_macros: Optional[ExcelHandler] = None
        self.corr_cols = CorrectionColumnInfo()

    def _update_progress(self, percent: int, message: str):
        """Обновление прогресса"""
        log.info(f"[{percent}%] {message}")
        if self.progress_callback:
            self.progress_callback(percent, message)

    @timing
    def process(self) -> ProcessingResult:
        """
        Основной метод обработки
        """
        result = ProcessingResult()

        try:
            self._update_progress(0, "Начало создания корректировки...")

            # 1. Открываем основной файл
            self._update_progress(5, "Открытие основного файла...")
            self.wb_macros = ExcelHandler(str(self.macros_file_path))
            self.wb_macros.open()

            # 2. Удаляем старый лист корректировки, если есть
            self._update_progress(10, "Подготовка листа корректировки...")
            if self.wb_macros.sheet_exists(SHEET_NAMES["correction"]):
                self.wb_macros.delete_sheet(SHEET_NAMES["correction"])

            # 3. Копируем шаблон
            self._update_progress(15, "Копирование шаблона...")
            if not self._copy_template():
                result.errors.append("Не удалось скопировать шаблон корректировки")
                return result

            # 4. Получаем информацию о колонках
            self._update_progress(20, "Анализ структуры колонок...")
            if not self._get_column_info():
                result.errors.append("Не удалось получить информацию о колонках")
                return result

            # 5. Собираем данные со всех расчётных листов
            self._update_progress(25, "Сбор данных с расчётных листов...")
            self._collect_data_from_calc_sheets()

            # 6. Заполняем общие реквизиты
            self._update_progress(75, "Заполнение общих реквизитов...")
            self._fill_common_fields()

            # 7. Удаляем нулевые и пустые строки
            self._update_progress(80, "Очистка нулевых записей...")
            self._cleanup_empty_rows()

            # 8. Форматирование
            self._update_progress(85, "Форматирование таблицы...")
            self._format_correction_sheet()

            # 9. Создаём контрольную формулу
            self._update_progress(90, "Создание контрольных формул...")
            self._create_control_formula()

            # 10. Перемещаем лист и скрываем шаблон
            self._update_progress(93, "Завершающие операции...")
            self._finalize_sheets()

            # 11. Сохраняем файл
            self._update_progress(95, "Сохранение файла...")
            self.wb_macros.save()

            self._update_progress(100, "Корректировка создана!")
            result.success = True
            result.message = 'Лист "Корректировка" загружен данными!'

        except Exception as e:
            log.exception(f"Ошибка обработки: {e}")
            result.errors.append(str(e))

        finally:
            if self.wb_macros:
                self.wb_macros.close()

        return result

    def _copy_template(self) -> bool:
        """Копирование шаблона корректировки"""
        ws_template = self.wb_macros.get_sheet(SHEET_NAMES["correction_template"])

        if not ws_template:
            log.error(f"Шаблон '{SHEET_NAMES['correction_template']}' не найден")
            return False

        # Делаем шаблон видимым
        ws_template.sheet_state = "visible"

        # Копируем лист
        ws_corr = self.wb_macros.copy_sheet(
            SHEET_NAMES["correction_template"], SHEET_NAMES["correction"]
        )

        log.info(f"Создан лист: {SHEET_NAMES['correction']}")
        return True

    def _get_column_info(self) -> bool:
        """Получение информации о колонках листа корректировки"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        if not ws_corr:
            return False

        # Колонки для поиска
        columns_to_find = [
            ("Балансовая Единица", "balance_unit"),
            ("Статья", "article"),
            ("ЦФО", "cfo"),
            ("Дебет", "debit"),
            ("Интерко2", "interco2"),
            ("Корректировки", "corrections"),
            ("Счет", "account"),
            ("Описание строки", "description"),
        ]

        arr_cols = search_arr_col(ws_corr, [c[0] for c in columns_to_find])

        for col_info in arr_cols:
            for search_name, attr_name in columns_to_find:
                if col_info["name"] == search_name:
                    setattr(self.corr_cols, attr_name, col_info["col_index"])
                    log.debug(f"Колонка '{search_name}': {col_info['col_letter']}")

        return self.corr_cols.balance_unit > 0

    def _collect_data_from_calc_sheets(self):
        """Сбор данных со всех расчётных листов"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        if not ws_corr:
            return

        helper_corr = SheetHelper(ws_corr)

        for iter_idx, sheet_name in enumerate(CALC_SHEETS, 1):
            progress = 25 + int((iter_idx / len(CALC_SHEETS)) * 50)
            self._update_progress(progress, f"Обработка: {sheet_name}")

            ws_calc = self.wb_macros.get_sheet(sheet_name)
            if not ws_calc:
                log.warning(f"Лист не найден: {sheet_name}")
                continue

            self._process_calc_sheet(ws_calc, ws_corr, sheet_name)

    def _process_calc_sheet(
        self, ws_calc: Worksheet, ws_corr: Worksheet, sheet_name: str
    ):
        """Обработка одного расчётного листа"""
        helper_calc = SheetHelper(ws_calc)
        helper_corr = SheetHelper(ws_corr)

        max_row_calc = helper_calc.max_row
        max_col_calc = helper_calc.max_col

        # Проверяем, есть ли данные
        if max_row_calc <= 6:
            log.debug(f"Нет данных на листе: {sheet_name}")
            return

        # Находим колонку "Сумма расходов с накопительным итогом" для определения конца сводной
        found_sum = helper_calc.find_value(
            "Сумма расходов с накопительным итогом", partial=False
        )
        if not found_sum:
            log.warning(f"Не найдена колонка суммы на листе: {sheet_name}")
            return

        end_pvt_col = found_sum[1]
        end_row_pvt = helper_calc.get_used_range_end(end_pvt_col) - 1

        # Определяем признаки листа
        is_cod = "ЦОД" in sheet_name
        is_vgo = "ИЗ ОТЧЕТА ВГО" in sheet_name

        # Обрабатываем два блока данных
        for i_rec in [1, 2]:
            current_row_corr = helper_corr.get_used_range_end(1)

            # Определяем диапазон для текущего блока
            if i_rec == 1:
                if is_vgo:
                    # Для ВГО первый блок начинается после сводной
                    start_row = end_row_pvt + 5
                    # Ищем блок Check для определения конца
                    found_check = helper_calc.find_value(
                        "Check",
                        partial=False,
                        start_row=end_row_pvt,
                        start_col=end_pvt_col + 1,
                    )
                    end_row = found_check[0] - 3 if found_check else max_row_calc
                else:
                    start_row = 6
                    end_row = end_row_pvt
            else:
                # Второй блок - после Check
                found_check = helper_calc.find_value(
                    "Check",
                    partial=False,
                    start_row=end_row_pvt,
                    start_col=end_pvt_col + 1,
                )
                if found_check:
                    start_row = found_check[0] + 3
                    end_row = helper_calc.get_used_range_end(end_pvt_col + 1)
                else:
                    continue

            # Колонки для копирования (из расчётного листа в корректировку)
            columns_map = [
                ("БЕ покупателя", self.corr_cols.balance_unit),
                (
                    "Статья КВ" if i_rec == 1 else "Статья операционная",
                    self.corr_cols.article,
                ),
                ("ЦФО КВ" if i_rec == 1 else "ЦФО операционное", self.corr_cols.cfo),
                ("Сумма расходов", self.corr_cols.debit),
                ("БЕ поставщика", self.corr_cols.interco2),
            ]

            for col_name, target_col in columns_map:
                if target_col <= 0:
                    continue

                found = helper_calc.find_value(
                    col_name,
                    partial=False,
                    start_row=start_row if i_rec == 1 and not is_vgo else end_row_pvt,
                    start_col=end_pvt_col + 1,
                )
                if not found:
                    # Пробуем найти в более широком диапазоне
                    found = helper_calc.find_value(
                        col_name, partial=False, start_row=1, start_col=end_pvt_col + 1
                    )

                if found:
                    source_col = found[1]
                    source_start_row = found[0] + 1
                    source_end_row = min(
                        end_row, helper_calc.get_used_range_end(source_col)
                    )

                    # Копируем данные
                    for row_offset in range(source_end_row - source_start_row + 1):
                        src_row = source_start_row + row_offset
                        dst_row = current_row_corr + 1 + row_offset

                        value = ws_calc.cell(row=src_row, column=source_col).value
                        ws_corr.cell(row=dst_row, column=target_col, value=value)

            # Заполняем колонку "Корректировки"
            final_row = helper_corr.get_used_range_end(self.corr_cols.balance_unit)

            correction_value = "CF16" if not is_cod else "CF16_1"
            for row in range(current_row_corr + 1, final_row + 1):
                ws_corr.cell(
                    row=row, column=self.corr_cols.corrections, value=correction_value
                )

            # Заполняем последнюю колонку названием листа
            max_col_corr = helper_corr.max_col
            for row in range(current_row_corr + 1, final_row + 1):
                ws_corr.cell(row=row, column=max_col_corr, value=sheet_name)

            # Для второго блока меняем местами БЕ и Интерко2
            if i_rec == 2:
                self._swap_be_interco(ws_corr, current_row_corr + 1, final_row)

        log.debug(f"Обработан лист: {sheet_name}")

    def _swap_be_interco(self, ws: Worksheet, start_row: int, end_row: int):
        """Замена местами Балансовой Единицы и Интерко2"""
        if self.corr_cols.balance_unit <= 0 or self.corr_cols.interco2 <= 0:
            return

        for row in range(start_row, end_row + 1):
            be_val = ws.cell(row=row, column=self.corr_cols.balance_unit).value
            interco_val = ws.cell(row=row, column=self.corr_cols.interco2).value

            # Меняем местами
            ws.cell(row=row, column=self.corr_cols.balance_unit, value=interco_val)
            ws.cell(row=row, column=self.corr_cols.balance_unit).number_format = "@"

            ws.cell(row=row, column=self.corr_cols.interco2, value=be_val)

    def _fill_common_fields(self):
        """Заполнение общих полей (Счёт, Описание строки)"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        if not ws_corr:
            return

        helper = SheetHelper(ws_corr)
        max_row = helper.get_used_range_end(self.corr_cols.balance_unit)

        # Заполняем Счёт
        if self.corr_cols.account > 0:
            for row in range(
                5, max_row + 1
            ):  # Начинаем с 5-й строки (после заголовков)
                ws_corr.cell(row=row, column=self.corr_cols.account, value="2600101")

        # Заполняем Описание строки
        if self.corr_cols.description > 0:
            description = "Корректировка инвестиционного оттока по ВГО договорам поставки в РТК от ДЗО"
            for row in range(5, max_row + 1):
                ws_corr.cell(
                    row=row, column=self.corr_cols.description, value=description
                )

        log.info(f"Заполнены общие поля для {max_row - 4} строк")

    def _cleanup_empty_rows(self):
        """Удаление нулевых и пустых строк"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        if not ws_corr:
            return

        helper = SheetHelper(ws_corr)

        # Удаляем строки с нулевым дебетом
        if self.corr_cols.debit > 0:
            rows_to_delete = []
            for row in range(helper.max_row, 4, -1):
                debit_val = ws_corr.cell(row=row, column=self.corr_cols.debit).value
                if debit_val == 0 or debit_val is None:
                    rows_to_delete.append(row)

            for row in rows_to_delete:
                ws_corr.delete_rows(row)

            log.debug(f"Удалено строк с нулевым дебетом: {len(rows_to_delete)}")

        # Удаляем строки с пустой Балансовой Единицей
        if self.corr_cols.balance_unit > 0:
            helper = SheetHelper(ws_corr)  # Обновляем после удаления
            rows_to_delete = []
            for row in range(helper.max_row, 4, -1):
                be_val = ws_corr.cell(row=row, column=self.corr_cols.balance_unit).value
                if not be_val:
                    rows_to_delete.append(row)

            for row in rows_to_delete:
                ws_corr.delete_rows(row)

            log.debug(f"Удалено строк с пустой БЕ: {len(rows_to_delete)}")

        # Приводим Интерко2 к текстовому формату
        if self.corr_cols.interco2 > 0:
            helper = SheetHelper(ws_corr)
            for row in range(5, helper.max_row + 1):
                cell = ws_corr.cell(row=row, column=self.corr_cols.interco2)
                cell.number_format = "@"

    def _format_correction_sheet(self):
        """Форматирование листа корректировки"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        if not ws_corr:
            return

        format_table(ws_corr, 2, 4)

    def _create_control_formula(self):
        """Создание контрольной формулы суммы"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])

        if not ws_corr or not ws_marja:
            return

        helper_marja = SheetHelper(ws_marja)

        # Ищем колонку с накопительным итогом на листе Маржа
        found = helper_marja.find_value(
            "Накопительный Итог_Расходы поставщика на выполнение работ по договору",
            partial=True,
        )
        if found:
            sum_row, sum_col = found
            col_letter = get_column_letter(sum_col)
            start_row = sum_row + 1
            end_row = helper_marja.get_used_range_end(sum_col)

            # Создаём формулу
            formula = f"=SUM('{SHEET_NAMES['marja']}'!{col_letter}{start_row}:{col_letter}{end_row})"
            ws_corr.cell(row=2, column=9, value=formula)

            log.info(f"Создана контрольная формула: {formula}")

    def _finalize_sheets(self):
        """Завершающие операции с листами"""
        ws_corr = self.wb_macros.get_sheet(SHEET_NAMES["correction"])
        ws_template = self.wb_macros.get_sheet(SHEET_NAMES["correction_template"])
        ws_start = self.wb_macros.get_sheet(SHEET_NAMES["start"])

        # Перемещаем лист корректировки после листа СТАРТ
        if ws_corr and ws_start:
            # В openpyxl нет прямого метода перемещения,
            # но можно изменить порядок через индексы
            wb = self.wb_macros.workbook
            sheets = wb._sheets

            # Находим индексы
            corr_idx = None
            start_idx = None
            for i, sheet in enumerate(sheets):
                if sheet.title == SHEET_NAMES["correction"]:
                    corr_idx = i
                elif sheet.title == SHEET_NAMES["start"]:
                    start_idx = i

            if corr_idx is not None and start_idx is not None:
                # Перемещаем лист
                sheet = sheets.pop(corr_idx)
                sheets.insert(start_idx + 1, sheet)

        # Скрываем шаблон
        if ws_template:
            ws_template.sheet_state = "hidden"
            log.debug("Шаблон корректировки скрыт")


def unload_corr(macros_file: str, progress_callback=None) -> ProcessingResult:
    """
    Основная функция для создания отчёта по корректировке

    Args:
        macros_file: Путь к основному файлу
        progress_callback: Callback для обновления прогресса

    Returns:
        ProcessingResult с результатом выполнения
    """
    processor = UnloadCorrProcessor(
        macros_file_path=macros_file, progress_callback=progress_callback
    )
    return processor.process()
