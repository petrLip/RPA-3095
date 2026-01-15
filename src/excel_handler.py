# -*- coding: utf-8 -*-
"""
Модуль работы с Excel файлами (без COM-объектов)
Поддерживает форматы: .xlsx, .xlsm, .xls, .xlsb
"""

import re
from pathlib import Path
from typing import Optional, List, Tuple, Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

try:
    from pyxlsb import open_workbook as open_xlsb
except ImportError:
    open_xlsb = None

from src.logger import log, timing, Timer


class ExcelHandler:
    """Класс для работы с Excel файлами"""

    def __init__(self, file_path: Optional[str] = None):
        self.file_path = Path(file_path) if file_path else None
        self.workbook: Optional[Workbook] = None
        self._is_xlsb = False

    @timing
    def open(
        self, file_path: Optional[str] = None, read_only: bool = False
    ) -> "ExcelHandler":
        """Открыть Excel файл"""
        if file_path:
            self.file_path = Path(file_path)

        if not self.file_path or not self.file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {self.file_path}")

        suffix = self.file_path.suffix.lower()
        log.info(f"Открытие файла: {self.file_path.name} (формат: {suffix})")

        if suffix == ".xlsb":
            self._is_xlsb = True
            self.workbook = self._read_xlsb_to_workbook()
        elif suffix in (".xlsx", ".xlsm"):
            # data_only=False чтобы сохранить формулы для последующего редактирования
            self.workbook = load_workbook(
                str(self.file_path), read_only=read_only, data_only=False
            )
        elif suffix == ".xls":
            # Читаем через pandas и конвертируем
            self.workbook = self._read_xls_to_workbook()
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {suffix}")

        log.info(f"Файл открыт. Листы: {self.workbook.sheetnames}")
        return self

    def _read_xlsb_to_workbook(self) -> Workbook:
        """Чтение xlsb файла и конвертация в Workbook"""
        if open_xlsb is None:
            raise ImportError(
                "Для работы с .xlsb файлами установите pyxlsb: pip install pyxlsb"
            )

        wb = Workbook()
        # Удаляем дефолтный лист
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        with Timer("Чтение XLSB файла"):
            with open_xlsb(str(self.file_path)) as xlsb_wb:
                for sheet_name in xlsb_wb.sheets:
                    log.debug(f"Чтение листа: {sheet_name}")
                    ws = wb.create_sheet(title=sheet_name)

                    with xlsb_wb.get_sheet(sheet_name) as xlsb_sheet:
                        for row_idx, row in enumerate(xlsb_sheet.rows(), 1):
                            for col_idx, cell in enumerate(row, 1):
                                ws.cell(row=row_idx, column=col_idx,
                                        value=cell.v)

        return wb

    def _read_xls_to_workbook(self) -> Workbook:
        """Чтение xls файла через pandas и конвертация в Workbook"""
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        with Timer("Чтение XLS файла"):
            xls = pd.ExcelFile(str(self.file_path), engine="xlrd")
            for sheet_name in xls.sheet_names:
                log.debug(f"Чтение листа: {sheet_name}")
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                ws = wb.create_sheet(title=sheet_name)

                for r_idx, row in enumerate(
                    dataframe_to_rows(df, index=False, header=False), 1
                ):
                    for c_idx, value in enumerate(row, 1):
                        if pd.notna(value):
                            ws.cell(row=r_idx, column=c_idx, value=value)

        return wb

    def create_workbook(self) -> "ExcelHandler":
        """Создать новую книгу"""
        self.workbook = Workbook()
        return self

    def save(self, file_path: Optional[str] = None, save_as_xlsx: bool = True):
        """
        Сохранить книгу

        Args:
            file_path: Путь для сохранения (если не указан, используется исходный путь)
            save_as_xlsx: Если True, сохраняет как .xlsx вместо .xlsm (по умолчанию True)
                         Это рекомендуется, так как макросы не сохраняются при работе через openpyxl
        """
        save_path = Path(file_path) if file_path else self.file_path
        if not save_path:
            raise ValueError("Путь для сохранения не указан")

        original_suffix = save_path.suffix.lower()
        original_path = save_path

        # xlsb нельзя сохранить в том же формате, сохраняем как xlsx
        if original_suffix == ".xlsb":
            save_path = save_path.with_suffix(".xlsx")
            log.warning(
                f"Формат xlsb не поддерживается для записи. Сохранение как: {save_path.name}"
            )
        # Если файл был .xlsm, но макросы не нужны, сохраняем как .xlsx
        elif original_suffix == ".xlsm" and save_as_xlsx:
            save_path = save_path.with_suffix(".xlsx")
            log.info(
                f"Файл сохранён как .xlsx (без макросов): {save_path.name}")
            log.info(
                f"Примечание: Макросы не сохраняются при работе через openpyxl. Для сохранения макросов используйте Excel."
            )

        with Timer(f"Сохранение файла: {save_path.name}"):
            # Убеждаемся, что файл сохраняется корректно
            try:
                # Сохраняем как обычный Excel файл (без макросов)
                self.workbook.save(str(save_path))
            except Exception as e:
                log.error(f"Ошибка при сохранении: {e}")
                # Пробуем сохранить как .xlsx, если была ошибка
                if original_suffix == ".xlsm" and save_path.suffix.lower() != ".xlsx":
                    save_path = original_path.with_suffix(".xlsx")
                    log.warning(
                        f"Повторная попытка сохранения как .xlsx: {save_path.name}"
                    )
                    self.workbook.save(str(save_path))
                else:
                    raise

        log.info(f"Файл успешно сохранён: {save_path}")

    def close(self):
        """Закрыть книгу"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def get_sheet(self, name: str) -> Optional[Worksheet]:
        """Получить лист по имени"""
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return None

    def get_or_create_sheet(self, name: str) -> Worksheet:
        """Получить или создать лист"""
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return self.workbook.create_sheet(title=name)

    def delete_sheet(self, name: str):
        """Удалить лист"""
        if name in self.workbook.sheetnames:
            del self.workbook[name]
            log.debug(f"Лист удалён: {name}")

    def sheet_exists(self, name: str) -> bool:
        """Проверить существование листа"""
        return name in self.workbook.sheetnames

    @property
    def sheet_names(self) -> List[str]:
        """Получить список имён листов"""
        return self.workbook.sheetnames if self.workbook else []

    def copy_sheet(self, source_name: str, target_name: str) -> Worksheet:
        """Копировать лист"""
        if source_name not in self.workbook.sheetnames:
            raise ValueError(f"Лист '{source_name}' не найден")

        source = self.workbook[source_name]
        target = self.workbook.copy_worksheet(source)
        target.title = target_name
        return target


class SheetHelper:
    """Вспомогательный класс для работы с листом Excel"""

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet

    @property
    def max_row(self) -> int:
        """Последняя строка с данными"""
        return self.ws.max_row or 1

    @property
    def max_col(self) -> int:
        """Последняя колонка с данными"""
        return self.ws.max_column or 1

    def get_used_range_end(self, column: int = 1) -> int:
        """Найти последнюю непустую строку в колонке"""
        for row in range(self.max_row, 0, -1):
            if self.ws.cell(row=row, column=column).value is not None:
                return row
        return 1

    def find_value(
        self,
        search_value: str,
        partial: bool = False,
        start_row: int = 1,
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
    ) -> Optional[Tuple[int, int]]:
        """
        Поиск значения в диапазоне
        Возвращает (row, col) или None
        """
        end_row = end_row or self.max_row
        end_col = end_col or self.max_col

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value)
                    if partial:
                        if search_value in cell_str:
                            return (row, col)
                    else:
                        if cell_str == search_value:
                            return (row, col)
        return None

    def find_all(
        self, search_value: str, partial: bool = False
    ) -> List[Tuple[int, int]]:
        """Найти все вхождения значения"""
        results = []
        for row in range(1, self.max_row + 1):
            for col in range(1, self.max_col + 1):
                cell_value = self.ws.cell(row=row, column=col).value
                if cell_value is not None:
                    cell_str = str(cell_value)
                    if partial and search_value in cell_str:
                        results.append((row, col))
                    elif not partial and cell_str == search_value:
                        results.append((row, col))
        return results

    def get_column_letter(self, col_index: int) -> str:
        """Получить букву колонки по индексу"""
        return get_column_letter(col_index)

    def get_column_index(self, col_letter: str) -> int:
        """Получить индекс колонки по букве"""
        return column_index_from_string(col_letter)

    def get_merged_cell_range(
        self, row: int, col: int
    ) -> Optional[Tuple[int, int, int, int]]:
        """Получить диапазон объединённой ячейки (min_row, min_col, max_row, max_col)"""
        # ReadOnlyWorksheet не поддерживает merged_cells
        if self._is_read_only():
            return None
        for merged_range in self.ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return (
                    merged_range.min_row,
                    merged_range.min_col,
                    merged_range.max_row,
                    merged_range.max_col,
                )
        return None

    def is_merged(self, row: int, col: int) -> bool:
        """Проверить, является ли ячейка объединённой"""
        # ReadOnlyWorksheet не поддерживает merged_cells
        if self._is_read_only():
            return False
        return self.get_merged_cell_range(row, col) is not None

    def copy_range(
        self,
        src_start_row: int,
        src_start_col: int,
        src_end_row: int,
        src_end_col: int,
        dest_row: int,
        dest_col: int,
        target_sheet: Optional[Worksheet] = None,
    ):
        """Копировать диапазон ячеек (только значения)"""
        target = target_sheet or self.ws

        for row_offset in range(src_end_row - src_start_row + 1):
            for col_offset in range(src_end_col - src_start_col + 1):
                src_cell = self.ws.cell(
                    row=src_start_row + row_offset, column=src_start_col + col_offset
                )
                dest_cell = target.cell(
                    row=dest_row + row_offset, column=dest_col + col_offset
                )
                dest_cell.value = src_cell.value

    def copy_column(
        self,
        src_col: int,
        dest_sheet: Worksheet,
        dest_col: int,
        start_row: int = 1,
        end_row: Optional[int] = None,
    ):
        """Копировать колонку на другой лист"""
        end_row = end_row or self.max_row

        for row in range(start_row, end_row + 1):
            value = self.ws.cell(row=row, column=src_col).value
            dest_sheet.cell(row=row, column=dest_col, value=value)

    def clear_sheet(self):
        """Очистить лист"""
        for row in self.ws.iter_rows():
            for cell in row:
                cell.value = None

    def delete_rows(self, start_row: int, count: int = 1):
        """Удалить строки"""
        self.ws.delete_rows(start_row, count)

    def delete_columns(self, start_col: int, count: int = 1):
        """Удалить колонки"""
        self.ws.delete_cols(start_col, count)

    def insert_rows(self, row: int, count: int = 1):
        """Вставить строки"""
        self.ws.insert_rows(row, count)

    def set_column_width(self, col: int, width: float):
        """Установить ширину колонки"""
        self.ws.column_dimensions[get_column_letter(col)].width = width

    def set_row_height(self, row: int, height: float):
        """Установить высоту строки"""
        self.ws.row_dimensions[row].height = height

    def apply_number_format(self, row: int, col: int, format_str: str):
        """Применить числовой формат"""
        self.ws.cell(row=row, column=col).number_format = format_str

    def get_data_as_list(
        self,
        start_row: int = 1,
        start_col: int = 1,
        end_row: Optional[int] = None,
        end_col: Optional[int] = None,
    ) -> List[List[Any]]:
        """Получить данные диапазона как список списков"""
        end_row = end_row or self.max_row
        end_col = end_col or self.max_col

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                row_data.append(self.ws.cell(row=row, column=col).value)
            data.append(row_data)
        return data

    def set_data_from_list(
        self, data: List[List[Any]], start_row: int = 1, start_col: int = 1
    ):
        """Записать данные из списка списков"""
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                self.ws.cell(
                    row=start_row + row_idx, column=start_col + col_idx, value=value
                )

    def fill_column_with_value(
        self, col: int, value: Any, start_row: int, end_row: int
    ):
        """Заполнить колонку значением"""
        for row in range(start_row, end_row + 1):
            self.ws.cell(row=row, column=col, value=value)

    def fill_column_with_formula(
        self, col: int, formula_template: str, start_row: int, end_row: int
    ):
        """
        Заполнить колонку формулой
        В formula_template используйте {row} для подстановки номера строки
        """
        for row in range(start_row, end_row + 1):
            formula = formula_template.replace("{row}", str(row))
            self.ws.cell(row=row, column=col, value=formula)

    def auto_filter(self, start_row: int, start_col: int, end_row: int, end_col: int):
        """Установить автофильтр"""
        range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        self.ws.auto_filter.ref = range_str

    def remove_auto_filter(self):
        """Удалить автофильтр"""
        # ReadOnlyWorksheet не поддерживает изменение auto_filter
        if hasattr(self.ws, "auto_filter") and not self._is_read_only():
            self.ws.auto_filter.ref = None

    def _is_read_only(self) -> bool:
        """Проверить, является ли лист только для чтения"""
        from openpyxl.worksheet._read_only import ReadOnlyWorksheet

        return isinstance(self.ws, ReadOnlyWorksheet)

    def unmerge_all_cells(self):
        """Разъединить все объединённые ячейки"""
        merged_ranges = list(self.ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            self.ws.unmerge_cells(str(merged_range))

    def show_all_rows(self):
        """Показать все скрытые строки"""
        # ReadOnlyWorksheet не поддерживает изменение row_dimensions
        if self._is_read_only():
            return
        for row in range(1, self.max_row + 1):
            self.ws.row_dimensions[row].hidden = False

    def show_all_columns(self):
        """Показать все скрытые колонки"""
        # ReadOnlyWorksheet не поддерживает изменение column_dimensions
        if self._is_read_only():
            return
        for col in range(1, self.max_col + 1):
            self.ws.column_dimensions[get_column_letter(col)].hidden = False

    def prepare_sheet(self):
        """Подготовить лист: снять фильтры, показать скрытые строки/колонки"""
        self.remove_auto_filter()
        self.show_all_rows()
        self.show_all_columns()
        # unmerge не делаем, т.к. может сломать логику


class TableFormatter:
    """Класс для форматирования таблиц"""

    # Стандартные стили границ
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    MEDIUM_BORDER = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    DOUBLE_BORDER = Border(
        left=Side(style="double"),
        right=Side(style="double"),
        top=Side(style="double"),
        bottom=Side(style="double"),
    )

    # Цвета заливки (RGB)
    FILL_COLORS = {
        "green_light": PatternFill(
            start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"
        ),
        "yellow_light": PatternFill(
            start_color="FFE699", end_color="FFE699", fill_type="solid"
        ),
        "blue_light": PatternFill(
            start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"
        ),
        "orange_light": PatternFill(
            start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"
        ),
        "red": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
    }

    def __init__(self, worksheet: Worksheet):
        self.ws = worksheet
        self.helper = SheetHelper(worksheet)

    def format_table(self, type_rpt: int, start_row: int = 1):
        """
        Форматирование таблицы в зависимости от типа отчёта
        type_rpt: 1 - МАРЖА, 2 - Корректировка, 3 - Табличная часть расчётов
        """
        max_row = self.helper.max_row
        max_col = self.helper.max_col

        if type_rpt == 1:
            self._format_marja(start_row, max_row, max_col)
        elif type_rpt == 2:
            self._format_correction(start_row, max_row, max_col)
        elif type_rpt == 3:
            self._format_calculation(start_row, max_row, max_col)

    def _format_marja(self, start_row: int, max_row: int, max_col: int):
        """Форматирование листа МАРЖА"""
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.MEDIUM_BORDER

        # Форматирование последних колонок
        for col in range(1, max_col + 1):
            header = self.ws.cell(row=start_row, column=col).value
            if header in [
                "Наименование покупателя",
                "*",
                "Источник данных для определения ЦФО и статьи",
                "Кластер",
            ]:
                self.ws.cell(row=start_row, column=col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                self.ws.cell(row=start_row, column=col).font = Font(
                    bold=True, size=9)

                if header == "*":
                    self.helper.set_column_width(col, 10)
                elif header == "Источник данных для определения ЦФО и статьи":
                    self.helper.set_column_width(col, 15)
                elif header == "Кластер":
                    self.helper.set_column_width(col, 8)

    def _format_correction(self, start_row: int, max_row: int, max_col: int):
        """Форматирование листа Корректировка"""
        for row in range(start_row, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = self.MEDIUM_BORDER

    def _format_calculation(self, start_row: int, max_row: int, max_col: int):
        """Форматирование табличной части расчётов

        Внешняя рамка каждой таблицы - двойная линия
        Внутренние границы ячеек - тонкая линия
        """
        # Находим все блоки таблиц и форматируем каждый отдельно
        table_blocks = self._find_table_blocks(start_row, max_row, 7, max_col)

        for block in table_blocks:
            self._format_single_table_block(
                block['start_row'],
                block['end_row'],
                block['start_col'],
                block['end_col']
            )

        # Заголовок
        for col in range(7, max_col + 1):
            cell = self.ws.cell(row=start_row, column=col)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        self.helper.set_row_height(start_row, 40)

        # Находим последнюю строку по колонке 7 для форматирования чисел
        last_row = self.helper.get_used_range_end(7)

        # Установка ширины колонок и форматов
        for col in range(7, max_col + 1):
            header = self.ws.cell(row=start_row, column=col).value

            if header in ["Сумма расходов", "Сумма расходов с накопительным итогом"]:
                for row in range(start_row, last_row + 1):
                    self.ws.cell(
                        row=row, column=col).number_format = "#,##0.00"
                self.helper.set_column_width(col, 30)

                if header == "Сумма расходов с накопительным итогом":
                    self.ws.cell(
                        row=start_row, column=col
                    ).value = (
                        "Сумма расходов с накопительным итогом расчетная для формул"
                    )

            elif header in ["БЕ + ЦФО", "БЕ поставщика", "ЦФО КВ", "Статья КВ"]:
                self.helper.set_column_width(col, 12)

            elif header in [
                "БЕ покупателя",
                "ЦФО операционное",
                "Статья операционная",
                "Статус анализа",
            ]:
                self.helper.set_column_width(col, 15)

            elif header == "№ инвест. Договора":
                self.helper.set_column_width(col, 50)

        # Цветовое оформление заголовка
        sheet_name = self.ws.title
        for col in range(7, max_col + 1):
            cell = self.ws.cell(row=start_row, column=col)

            if "БЮДЖЕТ" in sheet_name and "T2" not in sheet_name:
                cell.fill = self.FILL_COLORS["green_light"]
            elif "ОТЧЕТА ВГО" in sheet_name and "T2" not in sheet_name:
                cell.fill = self.FILL_COLORS["yellow_light"]
            elif "БЮДЖЕТ" in sheet_name and "T2" in sheet_name:
                cell.fill = self.FILL_COLORS["blue_light"]
            elif "ОТЧЕТА ВГО" in sheet_name and "T2" in sheet_name:
                cell.fill = self.FILL_COLORS["orange_light"]

        # Ширина служебных колонок
        self.helper.set_column_width(5, 5)
        self.helper.set_column_width(6, 5)

    def _find_table_blocks(
        self, start_row: int, max_row: int, start_col: int, max_col: int
    ) -> list:
        """
        Найти все блоки таблиц на листе.

        Блок определяется как последовательность непустых строк, начинающаяся
        с заголовка (БЕ + ЦФО, БЕ + Договор, БЕ поставщика и т.д.)

        Returns:
            Список словарей с границами каждого блока:
            [{'start_row': N, 'end_row': M, 'start_col': 7, 'end_col': K}, ...]
        """
        blocks = []
        current_block_start = None
        last_data_row = None

        # Заголовки, которые указывают на начало таблицы
        header_markers = [
            "БЕ + ЦФО", "БЕ + Договор", "БЕ поставщика",
        ]

        for row in range(start_row, max_row + 1):
            cell_value = self.ws.cell(row=row, column=start_col).value
            cell_str = str(cell_value) if cell_value else ""

            # Проверяем, является ли это строкой заголовка таблицы
            is_header = any(marker in cell_str for marker in header_markers)

            # Проверяем, есть ли данные в строке
            has_data = cell_value is not None and cell_str.strip() != ""

            # Пропускаем строки Check и пустые строки
            is_check_or_empty = cell_str.strip().lower() == "check" or not has_data

            if is_header and current_block_start is None:
                # Начало нового блока
                current_block_start = row
                last_data_row = row
            elif current_block_start is not None:
                if is_check_or_empty:
                    # Конец текущего блока (пустая строка или Check)
                    if last_data_row and last_data_row >= current_block_start:
                        # Находим последнюю колонку с данными в этом блоке
                        block_end_col = self._find_block_last_col(
                            current_block_start, last_data_row, start_col, max_col
                        )
                        blocks.append({
                            'start_row': current_block_start,
                            'end_row': last_data_row,
                            'start_col': start_col,
                            'end_col': block_end_col,
                        })
                    current_block_start = None
                    last_data_row = None
                elif has_data:
                    # Продолжение блока
                    last_data_row = row

        # Не забываем последний блок, если он не закрыт
        if current_block_start is not None and last_data_row:
            block_end_col = self._find_block_last_col(
                current_block_start, last_data_row, start_col, max_col
            )
            blocks.append({
                'start_row': current_block_start,
                'end_row': last_data_row,
                'start_col': start_col,
                'end_col': block_end_col,
            })

        return blocks

    def _find_block_last_col(
        self, start_row: int, end_row: int, start_col: int, max_col: int
    ) -> int:
        """Найти последнюю колонку с данными в блоке"""
        last_col = start_col

        # Проверяем строку заголовка (первую строку блока)
        for col in range(start_col, max_col + 1):
            cell_value = self.ws.cell(row=start_row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                last_col = col

        return last_col

    def _format_single_table_block(
        self, start_row: int, end_row: int, start_col: int, end_col: int
    ):
        """
        Форматирование одного блока таблицы.

        Внешняя рамка блока - двойная линия
        Внутренние границы ячеек - пунктирная линия
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = self.ws.cell(row=row, column=col)

                # Определяем тип границы для каждой стороны ячейки
                # Внешние края таблицы - двойная линия, внутренние - пунктирная
                left_style = "double" if col == start_col else "dashed"
                right_style = "double" if col == end_col else "dashed"
                top_style = "double" if row == start_row else "dashed"
                bottom_style = "double" if row == end_row else "dashed"

                cell.border = Border(
                    left=Side(style=left_style),
                    right=Side(style=right_style),
                    top=Side(style=top_style),
                    bottom=Side(style=bottom_style),
                )

    def highlight_cell(self, row: int, col: int, color: str = "red"):
        """Подсветить ячейку цветом"""
        if color in self.FILL_COLORS:
            self.ws.cell(row=row, column=col).fill = self.FILL_COLORS[color]

    def set_font_color(self, row: int, col: int, color: str):
        """Установить цвет шрифта"""
        color_map = {
            "red": "FF0000",
            "green": "008000",
            "blue": "0000FF",
        }
        rgb = color_map.get(color, color)
        self.ws.cell(row=row, column=col).font = Font(color=rgb)

    def set_bold(self, row: int, col: int, bold: bool = True):
        """Установить жирный шрифт"""
        current_font = self.ws.cell(row=row, column=col).font
        self.ws.cell(row=row, column=col).font = Font(
            bold=bold, size=current_font.size, color=current_font.color
        )


def read_excel_to_dataframe(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """Прочитать Excel файл в DataFrame"""
    path = Path(file_path)
    suffix = path.suffix.lower()

    with Timer(f"Чтение в DataFrame: {path.name}"):
        if suffix == ".xlsb":
            if open_xlsb is None:
                raise ImportError("Для .xlsb установите pyxlsb")

            with open_xlsb(str(path)) as wb:
                sheets = wb.sheets
                target_sheet = (
                    sheet_name if sheet_name and sheet_name in sheets else sheets[0]
                )

                with wb.get_sheet(target_sheet) as sheet:
                    data = []
                    for row in sheet.rows():
                        data.append([cell.v for cell in row])
                    return pd.DataFrame(data[1:], columns=data[0] if data else None)
        else:
            engine = "xlrd" if suffix == ".xls" else "openpyxl"
            return pd.read_excel(str(path), sheet_name=sheet_name, engine=engine)
