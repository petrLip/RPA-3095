# -*- coding: utf-8 -*-
"""
Вспомогательные функции (аналог модуля Others из VBA)
"""

from typing import List, Dict, Any, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from src.excel_handler import SheetHelper, TableFormatter
from src.logger import log, Timer


def search_arr_col(
    ws: Worksheet, arr_title: List[str], start_col: int = 0, end_col: int = 0
) -> List[Dict[str, Any]]:
    """
    Поиск нужных колонок для работы и заполнение информации о них
    Аналог процедуры SearchArrCol из VBA

    Args:
        ws: Рабочий лист
        arr_title: Список заголовков для поиска
        start_col: Начальная колонка диапазона поиска (0 = вся ширина)
        end_col: Конечная колонка диапазона поиска (0 = вся ширина)

    Returns:
        Список словарей с информацией о найденных колонках:
        [{'name': 'Название', 'col_index': 5, 'col_letter': 'E'}, ...]
    """
    helper = SheetHelper(ws)
    result = []

    max_row = helper.max_row
    max_col = helper.max_col

    s_col = start_col if start_col > 0 else 1
    e_col = end_col if end_col > 0 else max_col

    for title in arr_title:
        found = helper.find_value(
            title,
            partial=True,
            start_row=1,
            start_col=s_col,
            end_row=max_row,
            end_col=e_col,
        )
        if found:
            row, col = found
            result.append(
                {
                    "name": title,
                    "col_index": col,
                    "col_letter": get_column_letter(col),
                    "row": row,
                }
            )
            log.debug(f"Найдена колонка '{title}': {get_column_letter(col)}{row}")

    return result


def copy_paste_need_columns(
    source_ws: Worksheet, target_ws: Worksheet, arr_title: List[str], start_row: int = 1
) -> int:
    """
    Копирование необходимых колонок на временные листы
    Аналог процедуры CopyPastNeedColumns из VBA

    Args:
        source_ws: Исходный лист
        target_ws: Целевой лист
        arr_title: Список заголовков колонок для копирования
        start_row: Начальная строка на целевом листе

    Returns:
        Количество скопированных колонок
    """
    source_helper = SheetHelper(source_ws)
    target_helper = SheetHelper(target_ws)

    max_row = source_helper.max_row
    target_col = 1

    with Timer("Копирование колонок"):
        for title in arr_title:
            found = source_helper.find_value(title, partial=False)
            if found:
                row, col = found
                # Копируем колонку целиком
                source_helper.copy_column(
                    col, target_ws, target_col, start_row=row, end_row=max_row
                )
                # Перезаписываем заголовок (для точности)
                target_ws.cell(row=1, column=target_col, value=title)
                target_col += 1
                log.debug(f"Скопирована колонка '{title}'")

    return target_col - 1


def prepare_sheet_ext(ws: Worksheet):
    """
    Подготовка листа: отключение фильтров, показ скрытых строк/колонок
    Аналог процедуры PrepareShExt из VBA
    """
    helper = SheetHelper(ws)
    helper.prepare_sheet()
    log.debug(f"Лист '{ws.title}' подготовлен (фильтры сняты, строки/колонки показаны)")


def check_formulas(
    ws: Worksheet, start_row_range: int, type_rpt: int, sign: str = ""
) -> Dict[str, int]:
    """
    Создание формул для проверки расчётов на листах расчёты
    Аналог процедуры CheckFormuls из VBA

    Args:
        ws: Рабочий лист
        start_row_range: Начальная строка диапазона
        type_rpt: Тип отчёта (1-5)
        sign: Знак ("-" для отрицательных значений)

    Returns:
        Словарь с найденными колонками и позициями
    """
    helper = SheetHelper(ws)
    max_row = helper.max_row
    max_col = helper.max_col

    # Определяем количество итераций в зависимости от типа
    if type_rpt in [1, 4]:
        end_iter = 4
    elif type_rpt in [2, 3]:
        end_iter = 6
    elif type_rpt == 5:
        end_iter = 7
    else:
        end_iter = 4

    # Поиск ключевых колонок
    columns_info = {}

    # Поиск колонки "Сумма расходов с накопительным итогом"
    found = helper.find_value("Сумма расходов с накопительным итогом", partial=False)
    if found:
        row, col = found
        columns_info["end_pvt_blk"] = col
        end_pvt_blk = col
    else:
        log.warning("Не найдена колонка 'Сумма расходов с накопительным итогом'")
        return columns_info

    # Поиск остальных колонок
    search_cols = [
        ("БЕ поставщика", "be_sup"),
        ("БЕ покупателя", "be_buy"),
        ("ЦФО покупателя" if type_rpt != 4 else "Договор", "cfo_buy"),
    ]

    for search_name, key in search_cols:
        found = helper.find_value(search_name, partial=False, end_col=end_pvt_blk)
        if found:
            row, col = found
            columns_info[key] = col
            columns_info[f"{key}_row"] = row
            end_row_pvt = helper.get_used_range_end(col)
            columns_info["end_row_pvt"] = end_row_pvt
            columns_info["start_row_pvt"] = row + 1

    # Поиск колонок в правой части (мэппинг)
    search_mapping = [
        ("БЕ поставщика", "j_be"),
        ("%", "j_pr"),
        ("БЕ покупателя", "j_beb"),
    ]

    # Название 4-й колонки зависит от типа
    if type_rpt == 1:
        search_mapping.append(("ЦФО КВ", "j_cfo"))
    elif type_rpt in [2, 3, 5]:
        search_mapping.append(("ЦФО операционное", "j_cfo"))
    elif type_rpt == 4:
        search_mapping.append(("Договор", "j_cfo"))

    if type_rpt in [2, 3, 5]:
        search_mapping.append(("Сумма расходов с накопительным итогом", "j_sum_ras"))
        search_mapping.append(("Статья операционная", "j_stat"))

    if type_rpt == 5:
        search_mapping.append(("Договор", "j_dog"))

    for search_name, key in search_mapping:
        found = helper.find_value(
            search_name,
            partial=False,
            start_row=start_row_range,
            start_col=end_pvt_blk + 1,
        )
        if found:
            row, col = found
            columns_info[key] = col
            columns_info[f"{key}_row"] = row
            if key == "j_be":
                columns_info["end_row_j_data"] = helper.get_used_range_end(col)

    # Создание формулы SUMIFS
    _create_check_formula(ws, columns_info, type_rpt, sign, helper)

    return columns_info


def _create_check_formula(
    ws: Worksheet, cols: Dict[str, int], type_rpt: int, sign: str, helper: SheetHelper
):
    """Создание формулы проверки суммы"""
    if "j_be" not in cols or "end_pvt_blk" not in cols:
        log.warning("Недостаточно данных для создания формулы проверки")
        return

    start_row_pvt = cols.get("start_row_pvt", 2)
    end_row_pvt = cols.get("end_row_pvt", 100)
    end_pvt_blk = cols["end_pvt_blk"]
    j_be = cols.get("j_be", 1)
    j_pr = cols.get("j_pr", 1)
    i_be = cols.get("j_be_row", 1)
    end_row_j_data = cols.get("end_row_j_data", 100)
    max_col = helper.max_col

    be_sup_col = get_column_letter(cols.get("be_sup", 1))
    cfo_buy_col = get_column_letter(cols.get("cfo_buy", 1))
    be_buy_col = get_column_letter(cols.get("be_buy", 1))
    j_be_letter = get_column_letter(j_be)
    j_cfo_letter = get_column_letter(cols.get("j_cfo", 1))
    j_pr_letter = get_column_letter(j_pr)

    if type_rpt in [1, 4]:
        # Формула для типов 1 и 4
        ws.cell(row=i_be, column=max_col - 1, value="Сумма расходов")

        for row in range(i_be + 1, end_row_j_data + 1):
            formula = (
                f"={sign}SUMIFS("
                f"${get_column_letter(end_pvt_blk)}${start_row_pvt}:${get_column_letter(end_pvt_blk)}${end_row_pvt},"
                f"${be_sup_col}${start_row_pvt}:${be_sup_col}${end_row_pvt},{j_be_letter}{row},"
                f"${cfo_buy_col}${start_row_pvt}:${cfo_buy_col}${end_row_pvt},{j_cfo_letter}{row},"
                f"${be_buy_col}${start_row_pvt}:${be_buy_col}${end_row_pvt},{get_column_letter(max_col)}{row})"
                f"*{j_pr_letter}{row}%"
            )
            ws.cell(row=row, column=max_col - 1, value=formula)

        # Контрольная сумма
        ws.cell(
            row=i_be - 2,
            column=max_col,
            value=f"=SUM({get_column_letter(max_col - 1)}{i_be + 1}:{get_column_letter(max_col - 1)}{end_row_j_data})",
        )
        ws.cell(row=i_be - 2, column=max_col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=i_be - 2, column=max_col - 1, value="Check")
        ws.cell(row=i_be - 2, column=max_col - 1).font = openpyxl.styles.Font(
            color="FF0000", bold=True
        )

    elif type_rpt in [2, 3]:
        j_sum_ras = cols.get("j_sum_ras", 1)
        j_stat = cols.get("j_stat", 1)
        j_beb = cols.get("j_beb", 1)

        for row in range(i_be + 1, end_row_j_data + 1):
            formula = (
                f"={sign}SUMIFS("
                f"{get_column_letter(j_sum_ras)}{row}:{get_column_letter(j_sum_ras)}{row},"
                f"{j_be_letter}{row}:{j_be_letter}{row},{get_column_letter(max_col - 2 - j_be)}{row},"
                f"{get_column_letter(j_stat)}{row}:{get_column_letter(j_stat)}{row},{get_column_letter(max_col - 2 - j_stat)}{row},"
                f"{j_cfo_letter}{row}:{j_cfo_letter}{row},{get_column_letter(max_col - 2 - cols.get('j_cfo', 1))}{row},"
                f"{get_column_letter(j_beb)}{row}:{get_column_letter(j_beb)}{row},{get_column_letter(max_col)}{row})"
                f"*{j_pr_letter}{row}%"
            )
            ws.cell(row=row, column=max_col - 2, value=formula)

        # Контрольная сумма
        ws.cell(
            row=i_be - 2,
            column=max_col - 1,
            value=f"=SUM({get_column_letter(max_col - 2)}{i_be + 1}:{get_column_letter(max_col - 2)}{end_row_j_data})",
        )
        ws.cell(row=i_be - 2, column=max_col - 1).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=i_be - 2, column=max_col - 2, value="Check")
        ws.cell(row=i_be - 2, column=max_col - 2).font = openpyxl.styles.Font(
            color="FF0000", bold=True
        )

    log.debug(f"Формулы проверки созданы (type_rpt={type_rpt})")


import openpyxl.styles


def create_sheet_be_map(
    source_ws: Worksheet,
    start_col_account: int,
    type_rpt: int,
    arr_map: List[List[Any]],
    workbook,
) -> Optional[Worksheet]:
    """
    Создание листа с данными по Мэппинг + БЕ покупателя
    Аналог процедуры CreateSheetBEMap из VBA

    Args:
        source_ws: Исходный лист (расчёт)
        start_col_account: Начальная колонка для данных
        type_rpt: Тип отчёта (1, 2, 3)
        arr_map: Массив данных мэппинга
        workbook: Рабочая книга

    Returns:
        Созданный временный лист или None
    """
    with Timer("Создание листа БЕ_Мэппинг"):
        # Создаём временный лист
        if "БЕ_Мэппинг_БЕ_покупателя" in workbook.sheetnames:
            del workbook["БЕ_Мэппинг_БЕ_покупателя"]

        ws_map = workbook.create_sheet(title="БЕ_Мэппинг_БЕ_покупателя")

        # Заголовки в зависимости от типа
        if type_rpt == 1:
            headers = [
                "БЕ поставщика",
                "ЦФО операционное",
                "Статья операционная",
                "%",
                "Сумма расходов",
                "Сумма расходов с накопительным итогом",
                "БЕ покупателя",
            ]
        else:
            headers = [
                "БЕ поставщика",
                "ЦФО КВ" if type_rpt == 2 else "Договор",
                "ЦФО операционное",
                "Статья операционная",
                "%",
                "Сумма расходов",
                "Сумма расходов с накопительным итогом",
                "БЕ покупателя",
            ]

        for col, header in enumerate(headers, 1):
            ws_map.cell(row=1, column=col, value=header)

        # Находим ключевые колонки на исходном листе
        source_helper = SheetHelper(source_ws)

        found_sum = source_helper.find_value("Сумма расходов с накопительным итогом")
        if not found_sum:
            log.warning("Не найдена колонка суммы расходов")
            return ws_map

        end_pvt_col = found_sum[1]

        found_be_sup = source_helper.find_value("БЕ поставщика", end_col=end_pvt_col)
        found_be_buy = source_helper.find_value("БЕ покупателя", end_col=end_pvt_col)

        cfo_name = "ЦФО покупателя" if type_rpt != 3 else "Договор"
        found_cfo = source_helper.find_value(cfo_name, end_col=end_pvt_col)

        if not all([found_be_sup, found_be_buy, found_cfo]):
            log.warning("Не найдены все необходимые колонки")
            return ws_map

        start_row_pvt = found_be_sup[0] + 1
        end_row_pvt = source_helper.get_used_range_end(found_be_sup[1]) - 1

        # Заполняем данные
        target_row = 2

        for pvt_row in range(start_row_pvt, end_row_pvt + 1):
            be_sup_val = source_ws.cell(row=pvt_row, column=found_be_sup[1]).value
            be_buy_val = source_ws.cell(row=pvt_row, column=found_be_buy[1]).value
            cfo_val = source_ws.cell(row=pvt_row, column=found_cfo[1]).value
            sum_val = source_ws.cell(row=pvt_row, column=end_pvt_col).value

            # Ищем соответствия в arr_map
            for map_row in arr_map:
                if not map_row or not map_row[0]:
                    continue

                match = False
                if type_rpt == 1 and str(be_sup_val) == str(map_row[0]):
                    match = True
                elif (
                    type_rpt == 2
                    and str(be_sup_val) == str(map_row[1])
                    and str(cfo_val) == str(map_row[2])
                ):
                    match = True
                elif type_rpt == 3 and str(cfo_val) == str(map_row[2]):
                    match = True

                if match:
                    if type_rpt == 1:
                        ws_map.cell(row=target_row, column=1, value=str(be_sup_val))
                        ws_map.cell(
                            row=target_row,
                            column=2,
                            value=str(map_row[1]) if len(map_row) > 1 else "",
                        )
                        ws_map.cell(
                            row=target_row,
                            column=3,
                            value=map_row[2] if len(map_row) > 2 else "",
                        )
                        ws_map.cell(
                            row=target_row,
                            column=4,
                            value=map_row[3] if len(map_row) > 3 else "",
                        )
                        ws_map.cell(row=target_row, column=5, value="")
                        ws_map.cell(row=target_row, column=6, value=sum_val)
                        ws_map.cell(row=target_row, column=7, value=str(be_buy_val))
                    else:
                        ws_map.cell(row=target_row, column=1, value=str(be_sup_val))
                        ws_map.cell(
                            row=target_row,
                            column=2,
                            value=str(map_row[2]) if len(map_row) > 2 else "",
                        )
                        ws_map.cell(
                            row=target_row,
                            column=3,
                            value=str(map_row[3]) if len(map_row) > 3 else "",
                        )
                        ws_map.cell(
                            row=target_row,
                            column=4,
                            value=map_row[4] if len(map_row) > 4 else "",
                        )
                        ws_map.cell(
                            row=target_row,
                            column=5,
                            value=map_row[5] if len(map_row) > 5 else "",
                        )
                        ws_map.cell(row=target_row, column=6, value="")
                        ws_map.cell(row=target_row, column=7, value=sum_val)
                        ws_map.cell(row=target_row, column=8, value=str(be_buy_val))

                    target_row += 1

        log.info(f"Создан лист БЕ_Мэппинг с {target_row - 2} записями")
        return ws_map


def format_table(ws: Worksheet, type_rpt: int, start_row: int = 1):
    """
    Форматирование таблицы
    Аналог процедуры FormatTBL из VBA
    """
    formatter = TableFormatter(ws)
    formatter.format_table(type_rpt, start_row)
    log.debug(f"Лист '{ws.title}' отформатирован (type={type_rpt})")


def get_vlookup_formula(
    lookup_value_cell: str, table_range: str, col_index: int, approx_match: bool = False
) -> str:
    """
    Генерация формулы VLOOKUP (ВПР)

    Args:
        lookup_value_cell: Ячейка с искомым значением (напр. "A2")
        table_range: Диапазон таблицы (напр. "'Лист1'!A:D")
        col_index: Номер колонки для возврата
        approx_match: Приблизительное совпадение

    Returns:
        Строка с формулой VLOOKUP
    """
    match_type = "TRUE" if approx_match else "FALSE"
    return f"=VLOOKUP({lookup_value_cell},{table_range},{col_index},{match_type})"


def get_iferror_vlookup(
    lookup_value_cell: str, table_range: str, col_index: int, error_value: str = '""'
) -> str:
    """
    Генерация формулы IFERROR(VLOOKUP(...))
    """
    vlookup = f"VLOOKUP({lookup_value_cell},{table_range},{col_index},FALSE)"
    return f"=IFERROR({vlookup},{error_value})"


def get_sumifs_formula(sum_range: str, *criteria_pairs) -> str:
    """
    Генерация формулы SUMIFS

    Args:
        sum_range: Диапазон суммирования
        criteria_pairs: Пары (диапазон_критерия, значение_критерия)

    Returns:
        Строка с формулой SUMIFS
    """
    criteria_str = ",".join([f"{pair[0]},{pair[1]}" for pair in criteria_pairs])
    return f"=SUMIFS({sum_range},{criteria_str})"


def safe_str(value: Any) -> str:
    """Безопасное преобразование в строку"""
    if value is None:
        return ""
    return str(value)


def safe_float(value: Any, default: float = 0.0) -> float:
    """Безопасное преобразование в float"""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_int(value: Any, default: int = 0) -> int:
    """Безопасное преобразование в int"""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def pad_code(value: Any, length: int = 5) -> str:
    """Дополнить код до нужной длины нулями слева"""
    s = safe_str(value)
    if s.isdigit() and len(s) < length:
        return s.zfill(length)
    return s
