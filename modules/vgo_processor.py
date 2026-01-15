"""
Модуль обработки данных ВГО (эмуляция PivotTable через pandas)
Аналог блоков 1.8-1.11 из VBA макроса CreatePreviewData
"""

from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
import pandas as pd

from openpyxl.worksheet.worksheet import Worksheet

from modules.excel_handler import SheetHelper
from modules.helpers import safe_str, safe_float
from modules.logger import log


@dataclass
class VgoMappingResult:
    """Результат обработки мэппинга ВГО"""

    arr_err_be: List[Dict[str, Any]] = field(
        default_factory=list
    )  # Записи "Нет в мэппинге"
    arr_vgo_marja: List[Dict[str, Any]] = field(
        default_factory=list
    )  # Данные из МАРЖА для ВГО
    arr_rpt: List[Dict[str, Any]] = field(
        default_factory=list
    )  # Результаты поиска в сводной ВГО
    pivot_vgo_data: Optional[pd.DataFrame] = None  # Эмуляция сводной ВГО


class VgoProcessor:
    """
    Процессор для обработки данных ВГО
    Эмулирует логику VBA для листов РАСЧЕТ (ИЗ ОТЧЕТА ВГО)
    """

    START_COL_ACCOUNT = 7  # Стартовая колонка для начала данных на листах Расчёт

    def __init__(self, wb_macros, exclusions: List[str]):
        """
        Args:
            wb_macros: ExcelHandler основной книги
            exclusions: Список исключаемых статей
        """
        self.wb_macros = wb_macros
        self.exclusions = exclusions
        self.result = VgoMappingResult()

    def create_vgo_pivot_data(self) -> Optional[pd.DataFrame]:
        """
        Эмуляция сводной таблицы "Сводная_ВГО" через pandas
        Группировка: БЕ, ЦФО, Статья, Ссылка ID
        Значение: Сумма транзакции
        Фильтр: Account = "EXPENSE"

        Returns:
            DataFrame с агрегированными данными
        """
        ws_vgo = self.wb_macros.get_sheet("Данные_из_файла_по_ВГО")
        if not ws_vgo:
            log.warning("Лист 'Данные_из_файла_по_ВГО' не найден")
            return None

        helper = SheetHelper(ws_vgo)
        max_row = helper.max_row

        # Находим заголовки
        headers = {}
        for col in range(1, helper.max_col + 1):
            val = ws_vgo.cell(row=1, column=col).value
            if val:
                headers[str(val)] = col

        log.info(f"Заголовки листа ВГО: {list(headers.keys())}")

        # Колонки для извлечения
        col_be = headers.get("Балансовая единица (Entity)", 1)
        col_account = headers.get("Счет (Account)", 2)
        col_article = headers.get("Статья (IncomeExpenseItem)", 3)
        col_cfo = headers.get("ЦФО (CostCenter)", 4)
        col_ref_id = headers.get("Ссылка ID", 6)
        col_sum = headers.get("Сумма транзакции объекта", 7)

        # Читаем данные
        data = []
        for row in range(2, max_row + 1):
            account = safe_str(ws_vgo.cell(row=row, column=col_account).value)

            # Фильтр: только EXPENSE
            if account != "EXPENSE":
                continue

            data.append(
                {
                    "be": safe_str(ws_vgo.cell(row=row, column=col_be).value),
                    "cfo": safe_str(ws_vgo.cell(row=row, column=col_cfo).value),
                    "article": safe_str(ws_vgo.cell(row=row, column=col_article).value),
                    "ref_id": safe_str(
                        ws_vgo.cell(row=row, column=col_ref_id).value
                    ),  # № договора
                    "sum": safe_float(ws_vgo.cell(row=row, column=col_sum).value, 0),
                }
            )

        if not data:
            log.warning("Нет данных EXPENSE в файле ВГО")
            return None

        df = pd.DataFrame(data)

        # Группировка (эмуляция PivotTable)
        pivot = (
            df.groupby(["be", "cfo", "article", "ref_id"])
            .agg({"sum": "sum"})
            .reset_index()
        )

        log.info(f"Создана эмуляция сводной ВГО: {len(pivot)} записей")
        self.result.pivot_vgo_data = pivot

        return pivot

    def get_vgo_mapping_data(
        self, ws_map: Worksheet, is_t2: bool = False
    ) -> List[List[Any]]:
        """
        Получить данные из раздела мэппинга ВГО

        Args:
            ws_map: Лист мэппинга
            is_t2: True для T2 листов

        Returns:
            Список строк мэппинга
        """
        helper = SheetHelper(ws_map)

        section_name = (
            "Информация о ЦФО и статье из ВГО Т2"
            if is_t2
            else "Информация о ЦФО и статье из ВГО"
        )
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(f"Раздел мэппинга не найден: {section_name}")
            return []

        section_row, section_col = found

        # Определяем границы раздела (через MergeArea)
        merged = helper.get_merged_cell_range(section_row, section_col)
        end_col = merged[3] if merged else section_col + 10

        # Ищем строку с заголовками
        header_row = section_row + 1
        for check_row in range(section_row + 1, section_row + 5):
            first_val = ws_map.cell(row=check_row, column=section_col).value
            if first_val and "БЕ" in str(first_val):
                header_row = check_row
                break

        # Собираем данные
        result = []
        max_row = helper.max_row

        for row in range(header_row + 1, max_row + 1):
            row_data = []
            has_data = False
            for col in range(section_col, min(section_col + 15, end_col + 5)):
                val = ws_map.cell(row=row, column=col).value

                # Если это первая колонка (БЕ + ЦФО или БЕ + Договор) и значение - формула,
                # вычисляем её в Python
                if (
                    col == section_col
                    and val
                    and isinstance(val, str)
                    and val.startswith("=")
                ):
                    # Формула типа =AA5&AC5 - вычисляем сцепку
                    try:
                        # Парсим формулу =AA5&AC5
                        if "&" in val:
                            parts = val.split("&")
                            if len(parts) == 2:
                                # Извлекаем ссылки на ячейки (AA5, AC5)
                                ref1 = parts[0].replace("=", "").strip()
                                ref2 = parts[1].strip()

                                # Конвертируем ссылки в координаты
                                from openpyxl.utils import (
                                    coordinate_from_string,
                                    column_index_from_string,
                                )

                                col1, row1 = coordinate_from_string(ref1)
                                col2, row2 = coordinate_from_string(ref2)

                                # Читаем значения из этих ячеек
                                val1 = ws_map.cell(
                                    row=row1, column=column_index_from_string(col1)
                                ).value
                                val2 = ws_map.cell(
                                    row=row2, column=column_index_from_string(col2)
                                ).value

                                # Вычисляем сцепку
                                val = f"{safe_str(val1)}{safe_str(val2)}"
                    except Exception as e:
                        log.debug(
                            f"Не удалось вычислить формулу {val} в строке {row}: {e}"
                        )
                        # Оставляем исходное значение

                row_data.append(val)
                if val:
                    has_data = True

            if has_data:
                # Пропускаем заголовки
                first_val = safe_str(row_data[0])
                if first_val in ("БЕ + ЦФО", "БЕ + Договор", "БЕ поставщика"):
                    continue
                result.append(row_data)
            else:
                break  # Пустая строка - конец раздела

        log.info(f"Мэппинг ВГО '{section_name}': загружено {len(result)} записей")
        return result

    def find_err_be(
        self, ws_calc: Worksheet, vgo_mapping: List[List[Any]], is_t2: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Найти записи "Нет в мэппинге" (ArrErrBE)

        Args:
            ws_calc: Лист расчёта
            vgo_mapping: Данные мэппинга ВГО
            is_t2: True для T2 листов

        Returns:
            Список проблемных записей
        """
        helper = SheetHelper(ws_calc)

        # Создаём set ключей из мэппинга для быстрого поиска
        mapping_keys = set()
        for map_row in vgo_mapping:
            if map_row and map_row[0]:
                # Ключ = БЕ + ЦФО или БЕ + Договор
                key = safe_str(map_row[0])
                mapping_keys.add(key)

        log.info(f"Ключей в мэппинге ВГО: {len(mapping_keys)}")

        # Находим колонки на листе расчёта
        found_be = helper.find_value(
            "БЕ поставщика", partial=False, start_col=self.START_COL_ACCOUNT
        )

        if is_t2:
            found_key = helper.find_value(
                "Договор", partial=False, start_col=self.START_COL_ACCOUNT
            )
        else:
            found_key = helper.find_value(
                "ЦФО КВ", partial=False, start_col=self.START_COL_ACCOUNT
            )

        if not found_be or not found_key:
            log.warning("Не найдены колонки БЕ поставщика или ЦФО КВ/Договор")
            return []

        be_col = found_be[1]
        key_col = found_key[1]
        start_row = found_be[0] + 1
        end_row = helper.get_used_range_end(be_col)

        arr_err_be = []

        for row in range(start_row, end_row + 1):
            be_val = safe_str(ws_calc.cell(row=row, column=be_col).value)
            key_val = safe_str(ws_calc.cell(row=row, column=key_col).value)

            if not be_val:
                continue

            # Формируем ключ: БЕ + ЦФО
            lookup_key = f"{be_val}{key_val}"

            if lookup_key not in mapping_keys:
                arr_err_be.append(
                    {
                        "be": be_val,
                        "cfo": key_val,
                        "key": lookup_key,
                        "contract": "",  # Будет заполнен позже
                        "row": row,
                    }
                )

        log.info(f"Найдено записей 'Нет в мэппинге': {len(arr_err_be)}")
        self.result.arr_err_be = arr_err_be

        return arr_err_be

    def collect_vgo_marja_data(
        self, ws_marja: Worksheet, marja_cols: Any, start_row_tbl: int
    ) -> List[Dict[str, Any]]:
        """
        Собрать данные ArrVgoMarja - записи из МАРЖА где источник = "из отчета ВГО"

        Args:
            ws_marja: Лист МАРЖА
            marja_cols: MarjaColumnInfo с информацией о колонках
            start_row_tbl: Начальная строка таблицы

        Returns:
            Список записей для анализа ВГО
        """
        helper = SheetHelper(ws_marja)
        max_row = helper.max_row

        arr_vgo_marja = []

        for row in range(start_row_tbl + 1, max_row + 1):
            buyer_name = safe_str(
                ws_marja.cell(row=row, column=marja_cols.buyer_name).value
            )
            source_data = safe_str(
                ws_marja.cell(row=row, column=marja_cols.source_data).value
            )

            # Фильтр: источник = ВГО и покупатель != Т2
            if source_data == "из отчета ВГО" and buyer_name != "Т2":
                arr_vgo_marja.append(
                    {
                        "be": safe_str(
                            ws_marja.cell(
                                row=row, column=marja_cols.supplier_code
                            ).value
                        ),
                        "cfo": safe_str(
                            ws_marja.cell(row=row, column=marja_cols.cfo_buyer).value
                        ),
                        "contract": safe_str(
                            ws_marja.cell(
                                row=row, column=marja_cols.contract_number
                            ).value
                        ),
                    }
                )

        log.info(f"Собрано записей из МАРЖА для ВГО: {len(arr_vgo_marja)}")
        self.result.arr_vgo_marja = arr_vgo_marja

        return arr_vgo_marja

    def collect_vgo_marja_data_t2(
        self, ws_marja: Worksheet, marja_cols: Any, start_row_tbl: int
    ) -> List[Dict[str, Any]]:
        """
        Собрать данные из МАРЖА для ВГО T2 - записи с покупателем "Т2"

        Args:
            ws_marja: Лист МАРЖА
            marja_cols: MarjaColumnInfo с информацией о колонках
            start_row_tbl: Начальная строка таблицы

        Returns:
            Список записей для T2 с ЦФО покупателя (ЦФО КВ)
        """
        helper = SheetHelper(ws_marja)
        max_row = helper.max_row

        arr_vgo_marja_t2 = []

        for row in range(start_row_tbl + 1, max_row + 1):
            buyer_name = safe_str(
                ws_marja.cell(row=row, column=marja_cols.buyer_name).value
            )
            source_data = safe_str(
                ws_marja.cell(row=row, column=marja_cols.source_data).value
            )

            # Фильтр: источник = ВГО и покупатель = Т2
            if source_data == "из отчета ВГО" and buyer_name == "Т2":
                arr_vgo_marja_t2.append(
                    {
                        "be_supplier": safe_str(
                            ws_marja.cell(
                                row=row, column=marja_cols.supplier_code
                            ).value
                        ),
                        "cfo_buyer": safe_str(
                            ws_marja.cell(row=row, column=marja_cols.cfo_buyer).value
                        ),
                        "contract": safe_str(
                            ws_marja.cell(
                                row=row, column=marja_cols.contract_number
                            ).value
                        ),
                    }
                )

        log.info(f"Собрано записей из МАРЖА для ВГО T2: {len(arr_vgo_marja_t2)}")

        return arr_vgo_marja_t2

    def enrich_err_be_with_contracts(
        self, arr_err_be: List[Dict], arr_vgo_marja: List[Dict]
    ) -> List[Dict]:
        """
        Добавить номера договоров к проблемным записям из данных МАРЖА

        Args:
            arr_err_be: Проблемные записи
            arr_vgo_marja: Данные МАРЖА для ВГО

        Returns:
            Обогащённый список проблемных записей
        """
        for err in arr_err_be:
            if err["contract"]:
                continue  # Уже есть

            for marja in arr_vgo_marja:
                if err["be"] == marja["be"] and err["cfo"] == marja["cfo"]:
                    err["contract"] = marja["contract"]
                    break

        # Считаем сколько нашли договоров
        with_contracts = sum(1 for e in arr_err_be if e["contract"])
        log.info(
            f"Договора найдены для {with_contracts} из {len(arr_err_be)} проблемных записей"
        )

        return arr_err_be

    def search_in_vgo_pivot(
        self, arr_err_be: List[Dict], pivot_data: pd.DataFrame
    ) -> List[Dict]:
        """
        Поиск данных в сводной ВГО для каждой проблемной записи
        Эмуляция фильтрации PivotTable по Ссылка ID (№ договора)

        Args:
            arr_err_be: Проблемные записи с номерами договоров
            pivot_data: DataFrame эмуляции сводной ВГО

        Returns:
            Список результатов поиска (ArrRpt)
        """
        if pivot_data is None or pivot_data.empty:
            log.warning("Нет данных сводной ВГО для поиска")
            return []

        arr_rpt = []

        for err in arr_err_be:
            contract = err.get("contract", "")
            if not contract:
                # Нет договора - нет в выверке ВГО
                arr_rpt.append(
                    {
                        "be": err["be"],
                        "cfo": err["cfo"],
                        "cfo_oper": "Нет в выверке ВГО",
                        "article": "Нет в выверке ВГО",
                        "sum": 0,
                        "contract": contract,
                        "status": "Нет в выверке ВГО",
                    }
                )
                continue

            # Фильтруем по ref_id (№ договора)
            filtered = pivot_data[pivot_data["ref_id"] == contract].copy()

            if filtered.empty:
                arr_rpt.append(
                    {
                        "be": err["be"],
                        "cfo": err["cfo"],
                        "cfo_oper": "Нет в выверке ВГО",
                        "article": "Нет в выверке ВГО",
                        "sum": 0,
                        "contract": contract,
                        "status": "Нет в выверке ВГО",
                    }
                )
                continue

            # Удаляем статьи из списка исключений
            filtered = filtered[~filtered["article"].isin(self.exclusions)]

            if filtered.empty:
                arr_rpt.append(
                    {
                        "be": err["be"],
                        "cfo": err["cfo"],
                        "cfo_oper": "Нет в выверке ВГО",
                        "article": "Нет в выверке ВГО",
                        "sum": 0,
                        "contract": contract,
                        "status": "Нет в выверке ВГО",
                    }
                )
                continue

            # Сортировка по сумме и взятие топ 4-5 строк
            filtered = filtered.sort_values("sum", ascending=False).head(5)

            for _, row in filtered.iterrows():
                arr_rpt.append(
                    {
                        "be": err["be"],
                        "cfo": err["cfo"],
                        "cfo_oper": row["cfo"],
                        "article": row["article"],
                        "sum": row["sum"],
                        "contract": contract,
                        "status": "Есть в выверке ВГО",
                    }
                )

        log.info(
            f"Результаты поиска в ВГО: {len(arr_rpt)} записей "
            f"(Есть: {sum(1 for r in arr_rpt if r['status'] == 'Есть в выверке ВГО')})"
        )

        self.result.arr_rpt = arr_rpt
        return arr_rpt

    def autofill_vgo_mapping(
        self,
        ws_map: Worksheet,
        arr_rpt: List[Dict],
        ws_spr: Worksheet,
        is_t2: bool = False,
    ) -> int:
        """
        Автозаполнение мэппинга ВГО новыми строками из результатов поиска

        Args:
            ws_map: Лист мэппинга
            arr_rpt: Результаты поиска
            ws_spr: Справочник "Статья Операционая-Статья КВ"
            is_t2: True для T2 листов

        Returns:
            Количество добавленных строк
        """
        helper = SheetHelper(ws_map)

        # Находим раздел мэппинга
        section_name = (
            "Информация о ЦФО и статье из ВГО Т2"
            if is_t2
            else "Информация о ЦФО и статье из ВГО"
        )
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(f"Раздел мэппинга не найден для автозаполнения: {section_name}")
            return 0

        section_row, section_col = found

        # Находим колонки в разделе
        cols = {}
        for col_offset in range(0, 15):
            col = section_col + col_offset
            for check_row in range(section_row + 1, section_row + 4):
                val = ws_map.cell(row=check_row, column=col).value
                if val:
                    val_str = str(val)
                    if "БЕ + ЦФО" in val_str or "БЕ + Договор" in val_str:
                        cols["key"] = col
                    elif val_str == "БЕ поставщика":
                        cols["be"] = col
                    elif val_str == "ЦФО КВ" or val_str == "ЦФО операционное":
                        cols["cfo"] = col
                    elif val_str == "Статья операционная":
                        cols["article"] = col
                    elif val_str == "Статья PL":
                        cols["article_pl"] = col
                    elif val_str == "Расходы":
                        cols["expenses"] = col
                    elif val_str == "%":
                        cols["percent"] = col
                    break

        log.debug(f"Колонки мэппинга ВГО: {cols}")

        if not cols:
            log.warning("Не найдены колонки в разделе мэппинга ВГО")
            return 0

        # Создаём справочник статей
        article_dict = self._load_article_dict(ws_spr)

        # Находим последнюю строку в разделе
        key_col = cols.get("key", section_col)
        end_row = helper.get_used_range_end(key_col)

        added_count = 0

        for rpt in arr_rpt:
            if rpt["status"] != "Есть в выверке ВГО":
                continue

            new_row = end_row + 1 + added_count

            # БЕ + ЦФО (ключ)
            if "key" in cols:
                key_val = f"{rpt['be']}{rpt['cfo']}"
                ws_map.cell(row=new_row, column=cols["key"], value=key_val)

            # БЕ поставщика
            if "be" in cols:
                ws_map.cell(row=new_row, column=cols["be"], value=rpt["be"])
                ws_map.cell(row=new_row, column=cols["be"]).number_format = "@"

            # ЦФО КВ
            if "cfo" in cols:
                ws_map.cell(row=new_row, column=cols["cfo"], value=rpt["cfo"])
                ws_map.cell(row=new_row, column=cols["cfo"]).number_format = "@"

            # Статья операционная
            if "article" in cols:
                ws_map.cell(row=new_row, column=cols["article"], value=rpt["article"])

            # Статья PL (через VLOOKUP по справочнику)
            if "article_pl" in cols:
                article_pl = article_dict.get(rpt["article"], "Статья не найдена")
                ws_map.cell(row=new_row, column=cols["article_pl"], value=article_pl)

            # Расходы
            if "expenses" in cols and rpt["sum"]:
                ws_map.cell(row=new_row, column=cols["expenses"], value=rpt["sum"])

            # % - будет рассчитан позже или 100 по умолчанию
            if "percent" in cols:
                ws_map.cell(row=new_row, column=cols["percent"], value=100)

            added_count += 1

        log.info(f"Добавлено {added_count} новых строк в мэппинг ВГО")
        return added_count

    def _load_article_dict(self, ws_spr: Worksheet) -> Dict[str, str]:
        """Загрузить справочник статей"""
        if not ws_spr:
            return {}

        helper = SheetHelper(ws_spr)
        result = {}

        for row in range(2, helper.max_row + 1):
            key = safe_str(ws_spr.cell(row=row, column=1).value)
            val = safe_str(ws_spr.cell(row=row, column=2).value)
            if key:
                result[key] = val

        log.debug(f"Загружен справочник статей: {len(result)} записей")
        return result

    def evaluate_vgo_t2_mapping_formulas(self, ws_map: Worksheet):
        """
        Вычисляет формулы "БЕ + Договор" в мэппинге ВГО Т2 и вставляет вычисленные значения

        Args:
            ws_map: Лист мэппинга
        """
        helper = SheetHelper(ws_map)

        # Находим раздел мэппинга ВГО Т2
        section_name = "Информация о ЦФО и статье из ВГО Т2"
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.debug(f"Секция '{section_name}' не найдена для вычисления формул")
            return

        log.debug(f"Найдена секция '{section_name}' в ({found[0]}, {found[1]})")

        section_row, section_col = found

        # Ищем заголовки
        header_row = section_row + 1
        for check_row in range(section_row + 1, section_row + 5):
            first_val = ws_map.cell(row=check_row, column=section_col).value
            if first_val and "БЕ" in str(first_val):
                header_row = check_row
                break

        # Находим колонки: БЕ + Договор (Z), БЕ поставщика (AA), Договор (AC)
        key_col = None  # Z = БЕ + Договор
        be_col = None  # AA = БЕ поставщика
        dog_col = None  # AC = Договор

        for col_offset in range(0, 10):
            col = section_col + col_offset
            for check_row in range(header_row, header_row + 2):
                val = ws_map.cell(row=check_row, column=col).value
                if val:
                    val_str = str(val)
                    if "БЕ + Договор" in val_str:
                        key_col = col
                    elif val_str == "БЕ поставщика":
                        be_col = col
                    elif val_str == "Договор":
                        dog_col = col

        if not key_col or not be_col or not dog_col:
            log.warning(
                f"Не найдены колонки для вычисления формул в мэппинге ВГО Т2: key_col={key_col}, be_col={be_col}, dog_col={dog_col}"
            )
            return

        log.debug(
            f"Колонки найдены: key_col={key_col} (Z), be_col={be_col} (AA), dog_col={dog_col} (AC)"
        )

        # Вычисляем формулы в колонке "БЕ + Договор"
        max_row = helper.max_row
        evaluated_count = 0

        for row in range(header_row + 1, max_row + 1):
            # Проверяем, есть ли формула в колонке "БЕ + Договор"
            val = ws_map.cell(row=row, column=key_col).value

            if val and isinstance(val, str) and val.startswith("="):
                # Формула типа =AA5&AC5 - вычисляем
                try:
                    if "&" in val:
                        parts = val.split("&")
                        if len(parts) == 2:
                            # Извлекаем ссылки на ячейки
                            from openpyxl.utils import (
                                coordinate_from_string,
                                column_index_from_string,
                            )

                            ref1 = parts[0].replace("=", "").strip()
                            ref2 = parts[1].strip()

                            col1, row1 = coordinate_from_string(ref1)
                            col2, row2 = coordinate_from_string(ref2)

                            # Читаем значения
                            val1 = ws_map.cell(
                                row=row1, column=column_index_from_string(col1)
                            ).value
                            val2 = ws_map.cell(
                                row=row2, column=column_index_from_string(col2)
                            ).value

                            # Вычисляем и вставляем значение вместо формулы
                            calculated = f"{safe_str(val1)}{safe_str(val2)}"
                            ws_map.cell(row=row, column=key_col).value = calculated
                            evaluated_count += 1
                except Exception as e:
                    log.debug(f"Не удалось вычислить формулу {val} в строке {row}: {e}")

        if evaluated_count > 0:
            log.info(
                f"Вычислено {evaluated_count} формул 'БЕ + Договор' в мэппинге ВГО Т2"
            )

    def update_analysis_status(
        self, ws_calc: Worksheet, arr_err_be: List[Dict], arr_rpt: List[Dict]
    ):
        """
        Обновить статус анализа и № договора на листе расчёта

        Args:
            ws_calc: Лист расчёта
            arr_err_be: Проблемные записи
            arr_rpt: Результаты поиска
        """
        helper = SheetHelper(ws_calc)

        # Находим колонки
        found_status = helper.find_value(
            "Статус анализа", partial=False, start_col=self.START_COL_ACCOUNT
        )
        found_contract = helper.find_value(
            "№ инвест. Договора", partial=False, start_col=self.START_COL_ACCOUNT
        )

        if not found_status:
            log.warning("Колонка 'Статус анализа' не найдена")
            return

        status_col = found_status[1]
        contract_col = found_contract[1] if found_contract else None

        # Создаём словарь статусов
        # По умолчанию все записи имеют статус "ОК" (установлен при записи)
        # Обновляем только если найден в выверке ВГО - ставим договор
        contract_dict = {}

        for rpt in arr_rpt:
            key = f"{rpt['be']}_{rpt['cfo']}"
            if rpt["status"] == "Есть в выверке ВГО":
                contract_dict[key] = rpt.get("contract", "")
            # Если не найден в выверке - оставляем статус "ОК" по умолчанию

        # Обновляем только договора для записей найденных в выверке
        updated = 0
        for err in arr_err_be:
            row = err["row"]
            key = f"{err['be']}_{err['cfo']}"

            # Договор из результатов поиска
            if key in contract_dict and contract_col:
                ws_calc.cell(row=row, column=contract_col, value=contract_dict[key])
                updated += 1
            # Договор из МАРЖА
            elif contract_col and err.get("contract"):
                ws_calc.cell(row=row, column=contract_col, value=err["contract"])

        log.info(
            f"Обновлены договора для {updated} записей (статусы оставлены 'ОК' по умолчанию)"
        )

    def create_vgo_second_block(
        self,
        ws_calc: Worksheet,
        pivot_data: Dict,
        vgo_mapping: List[List[Any]],
        is_t2: bool = False,
    ):
        """
        Создание второго блока (CAPEX) для обычных листов ВГО (не T2)

        Структура в эталоне:
        - Check перед блоком (сумма со знаком минус)
        - Заголовки CAPEX: БЕ + ЦФО, БЕ поставщика, ЦФО КВ, Статья КВ, %, Сумма расходов, БЕ покупателя
        - Данные

        Args:
            ws_calc: Лист расчёта
            pivot_data: Данные из сводной таблицы
            vgo_mapping: Данные мэппинга ВГО
            is_t2: True для T2 листов
        """
        helper = SheetHelper(ws_calc)

        # Находим конец первого блока
        end_row_first = helper.get_used_range_end(self.START_COL_ACCOUNT)

        # Фильтруем pivot_data - исключаем "Общий итог"
        filtered_pivot = {
            k: v
            for k, v in pivot_data.items()
            if safe_str(v.get("be_supplier", "")) != "Общий итог"
        }

        # Check перед вторым блоком - вычисляем сумму
        check_row = end_row_first + 3
        total_sum = sum(safe_float(v.get("sum", 0), 0) for v in filtered_pivot.values())
        ws_calc.cell(row=check_row, column=self.START_COL_ACCOUNT + 5, value="Check")
        ws_calc.cell(row=check_row, column=self.START_COL_ACCOUNT + 6, value=-total_sum)

        # Заголовки второго блока (CAPEX для обычных ВГО)
        current_row = check_row + 2

        # Для обычных ВГО - заголовки CAPEX
        headers = [
            "БЕ + ЦФО",
            "БЕ поставщика",
            "ЦФО КВ",
            "Статья КВ",
            "%",
            "Сумма расходов",
            "БЕ покупателя",
        ]

        # Записываем заголовки
        for col_offset, header in enumerate(headers):
            ws_calc.cell(
                row=current_row,
                column=self.START_COL_ACCOUNT + col_offset,
                value=header,
            )

        # Заполняем данные CAPEX
        # Структура: БЕ + ЦФО, БЕ поставщика, ЦФО КВ, Статья КВ, %, Сумма расходов, БЕ покупателя
        data_row = current_row + 1

        for key, pv_values in filtered_pivot.items():
            be_supplier = safe_str(pv_values["be_supplier"])
            cfo = safe_str(pv_values["cfo"])

            base_sum = safe_float(pv_values.get("sum", 0), 0)
            percent_val = 100
            calculated_sum = -base_sum * percent_val / 100  # Отрицательное для CAPEX

            # БЕ + ЦФО
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT, value=f"{be_supplier}{cfo}"
            )
            # БЕ поставщика
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 1, value=be_supplier
            )
            # ЦФО КВ
            ws_calc.cell(row=data_row, column=self.START_COL_ACCOUNT + 2, value=cfo)
            # Статья КВ
            ws_calc.cell(row=data_row, column=self.START_COL_ACCOUNT + 3, value="PI02")
            # %
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 4, value=percent_val
            )
            # Сумма расходов (отрицательная)
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 5, value=calculated_sum
            )
            # БЕ покупателя
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 6,
                value=pv_values["be_buyer"],
            )

            data_row += 1

        capex_end_row = data_row - 1
        log.info(
            f"Создан второй блок ВГО (CAPEX): {data_row - current_row - 1} записей, начиная со строки {current_row}"
        )

        # ==================== ТРЕТИЙ БЛОК (OPEX) ====================
        # Check перед OPEX - будет заполнен после создания OPEX блока
        check_row_opex = capex_end_row + 3
        ws_calc.cell(
            row=check_row_opex, column=self.START_COL_ACCOUNT + 5, value="Check"
        )
        # Значение Check будет вычислено после создания OPEX блока

        # Заголовки OPEX
        opex_header_row = check_row_opex + 2
        headers_opex = [
            "БЕ + ЦФО",
            "БЕ поставщика",
            "ЦФО КВ",
            "ЦФО операционное",
            "Статья операционная",
            "%",
            "Сумма расходов",
            "Сумма расходов с накопительным итогом расчетная для формул",
            "БЕ покупателя",
        ]

        for col_offset, header in enumerate(headers_opex):
            ws_calc.cell(
                row=opex_header_row,
                column=self.START_COL_ACCOUNT + col_offset,
                value=header,
            )

        # Очищаем область данных OPEX (удаляем старые данные/формулы)
        # Максимум 200 строк для очистки
        for row in range(opex_header_row + 1, opex_header_row + 200):
            for col in range(self.START_COL_ACCOUNT, self.START_COL_ACCOUNT + 10):
                ws_calc.cell(row=row, column=col).value = None

        # Создаём словарь мэппинга для OPEX
        # Исключаем дубликаты записей в самом мэппинге
        mapping_dict = {}
        seen_mapping_rows = set()  # Для исключения полных дубликатов в мэппинге

        for map_row in vgo_mapping:
            if not map_row or len(map_row) < 2:
                continue
            # Создаём ключ из БЕ + ЦФО КВ
            key_val = safe_str(map_row[0])
            if key_val.startswith("="):
                # Если формула - составляем из отдельных колонок
                be = safe_str(map_row[1])
                cfo = safe_str(map_row[2]) if len(map_row) > 2 else ""
                key_val = f"{be}{cfo}"

            if key_val and key_val not in ("БЕ + ЦФО", "БЕ поставщика"):
                # Создаём уникальный идентификатор записи мэппинга
                # для исключения полных дубликатов (одинаковые BE, CFO_KV, CFO_OPER, STAT, %)
                row_id = (
                    safe_str(map_row[1]),  # BE
                    safe_str(map_row[2]) if len(map_row) > 2 else "",  # CFO_KV
                    safe_str(map_row[3]) if len(map_row) > 3 else "",  # CFO_OPER
                    safe_str(map_row[5]) if len(map_row) > 5 else "",  # STAT_OPER
                    safe_str(map_row[7]) if len(map_row) > 7 else "",  # %
                )

                if row_id in seen_mapping_rows:
                    continue  # Пропускаем дубликат
                seen_mapping_rows.add(row_id)

                if key_val not in mapping_dict:
                    mapping_dict[key_val] = []
                mapping_dict[key_val].append(map_row)

        log.info(f"Мэппинг ВГО для OPEX блока: {len(mapping_dict)} ключей")

        # Загружаем справочник статей для вычисления формул VLOOKUP
        ws_spr = self.wb_macros.get_sheet("Статья Операционая-Статья КВ")
        article_dict = self._load_article_dict(ws_spr)

        # Предвычисляем суммы расходов по (БЕ, ЦФО КВ) для формулы процента
        # Формула: =IF(W="",100,(W*100)/SUMIFS(W:W,R:R,R,S:S,S))
        expense_sums = {}
        for map_row in vgo_mapping:
            if not map_row or len(map_row) < 7:
                continue
            be = safe_str(map_row[1])
            cfo_kv = safe_str(map_row[2]) if len(map_row) > 2 else ""
            expense = safe_float(map_row[6], 0) if len(map_row) > 6 else 0
            key = (be, cfo_kv)
            expense_sums[key] = expense_sums.get(key, 0) + expense

        # Заполняем данные OPEX
        # В VBA: для КАЖДОЙ строки pivot (включая разные be_buyer) создаются записи OPEX
        # Ключ pivot_data включает be_buyer, поэтому каждая комбинация (be, cfo, be_buyer) обрабатывается отдельно

        # Сначала собираем все строки, потом сортируем как в VBA
        opex_rows_to_write = []

        for key, pv_values in filtered_pivot.items():
            be_supplier = safe_str(pv_values["be_supplier"])
            cfo = safe_str(pv_values["cfo"])
            be_buyer = safe_str(pv_values.get("be_buyer", ""))
            lookup_key = f"{be_supplier}{cfo}"

            if lookup_key in mapping_dict:
                matched_rows = mapping_dict[lookup_key]

                for map_row in matched_rows:
                    # Структура мэппинга ВГО (секция Q:X):
                    # [0]=БЕ+ЦФО (Q), [1]=БЕ (R), [2]=ЦФО КВ (S), [3]=ЦФО опер (T),
                    # [4]=Статья PL (U), [5]=Статья опер (V), [6]=Расходы (W), [7]=% (X)
                    cfo_oper = safe_str(map_row[3]) if len(map_row) > 3 else ""
                    stat_oper = safe_str(map_row[5]) if len(map_row) > 5 else ""
                    stat_pl = (
                        safe_str(map_row[4]) if len(map_row) > 4 else ""
                    )  # Статья PL для VLOOKUP
                    percent_raw = safe_str(map_row[7]) if len(map_row) > 7 else "100"

                    # Если stat_oper - формула VLOOKUP, вычисляем через справочник
                    if safe_str(stat_oper).startswith("="):
                        # Формула типа: =IFERROR(VLOOKUP(U117,'Статья Операционая-Статья КВ'!A:B,2,0),"Статья не найдена")
                        # Используем Статью PL для поиска в справочнике
                        if stat_pl and not safe_str(stat_pl).startswith("="):
                            stat_oper = article_dict.get(stat_pl, "")
                            if not stat_oper:
                                log.debug(
                                    f"Статья PL '{stat_pl}' не найдена в справочнике"
                                )
                        else:
                            stat_oper = ""

                    # Если percent_val - формула, вычисляем через SUMIFS
                    if safe_str(percent_raw).startswith("="):
                        # Формула: =IF(W="",100,(W*100)/SUMIFS(W:W,R:R,R,S:S,S))
                        be = safe_str(map_row[1])
                        cfo_kv = safe_str(map_row[2]) if len(map_row) > 2 else ""
                        expense = safe_float(map_row[6], 0) if len(map_row) > 6 else 0
                        sum_expense = expense_sums.get((be, cfo_kv), 0)
                        if expense and sum_expense:
                            percent_val = (expense * 100) / sum_expense
                        else:
                            percent_val = 100.0
                    else:
                        percent_val = safe_float(percent_raw, 100)

                    # Пропускаем если нет данных или формула
                    skip_reason = None
                    if not cfo_oper:
                        skip_reason = "no cfo_oper"
                    elif not stat_oper:
                        skip_reason = "no stat_oper"
                    elif safe_str(cfo_oper).startswith("="):
                        skip_reason = f"cfo_oper is formula: {cfo_oper}"
                    elif stat_oper == "Статья не найдена":
                        skip_reason = "stat_oper = Статья не найдена"

                    if skip_reason:
                        continue

                    base_sum = safe_float(pv_values.get("sum", 0), 0)
                    calculated_sum = base_sum * percent_val / 100

                    opex_rows_to_write.append(
                        {
                            "be_cfo": f"{be_supplier}{cfo}",
                            "be_supplier": be_supplier,
                            "cfo": cfo,
                            "cfo_oper": cfo_oper,
                            "stat_oper": stat_oper,
                            "percent_val": percent_val,
                            "calculated_sum": calculated_sum,
                            "base_sum": base_sum,
                            "be_buyer": pv_values["be_buyer"],
                        }
                    )

        # Сортируем по base_sum (убывание) как в VBA через PivotTable
        opex_rows_to_write.sort(key=lambda x: -abs(x["base_sum"]))

        # Записываем отсортированные данные
        data_row = opex_header_row + 1
        for row_data in opex_rows_to_write:
            # БЕ + ЦФО
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT, value=row_data["be_cfo"]
            )
            # БЕ поставщика
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 1,
                value=row_data["be_supplier"],
            )
            # ЦФО КВ
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 2, value=row_data["cfo"]
            )
            # ЦФО операционное
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 3,
                value=row_data["cfo_oper"],
            )
            # Статья операционная
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 4,
                value=row_data["stat_oper"],
            )
            # %
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 5,
                value=row_data["percent_val"],
            )
            # Сумма расходов
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 6,
                value=row_data["calculated_sum"],
            )
            # Сумма расходов с накопительным итогом
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 7,
                value=row_data["base_sum"],
            )
            # БЕ покупателя
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 8,
                value=row_data["be_buyer"],
            )

            data_row += 1

        # Вычисляем сумму OPEX блока для Check
        opex_sum = 0
        for row in range(opex_header_row + 1, data_row):
            val = ws_calc.cell(row=row, column=self.START_COL_ACCOUNT + 6).value
            if val:
                opex_sum += safe_float(val, 0)

        # Записываем Check (сумма M колонки OPEX блока)
        ws_calc.cell(
            row=check_row_opex, column=self.START_COL_ACCOUNT + 6, value=opex_sum
        )

        log.info(
            f"Создан третий блок ВГО (OPEX): {data_row - opex_header_row - 1} записей, начиная со строки {opex_header_row}"
        )

    def create_vgo_t2_blocks(
        self,
        ws_calc: Worksheet,
        pivot_data: Dict,
        vgo_mapping: List[List[Any]],
        arr_vgo_marja: List[Dict[str, Any]],
    ):
        """
        Создание блоков CAPEX и OPEX для листов ВГО T2

        Структура эталона:
        - Строки 5-14: Первый блок (Сводная + Статус анализа)
        - Строка 17: Check (отрицательная сумма)
        - Строки 19-28: Второй блок (CAPEX): ЦФО КВ, Статья КВ
        - Строка 31: Check (положительная сумма)
        - Строки 33-42: Третий блок (OPEX): ЦФО операционное, Статья операционная

        Args:
            arr_vgo_marja: Данные из МАРЖА (для получения ЦФО КВ = ЦФО покупателя)
        """
        helper = SheetHelper(ws_calc)

        # Находим конец первого блока
        end_row_first = helper.get_used_range_end(self.START_COL_ACCOUNT)

        # Фильтруем pivot_data - исключаем "Общий итог"
        filtered_pivot = {
            k: v
            for k, v in pivot_data.items()
            if safe_str(v.get("be_supplier", "")) != "Общий итог"
        }

        # Создаём словарь ЦФО КВ из данных МАРЖА
        # Ключ = (БЕ поставщика + Договор), значение = ЦФО покупателя
        cfo_kv_dict = {}
        for marja_row in arr_vgo_marja:
            be_sup = safe_str(marja_row.get("be_supplier", ""))
            contract = safe_str(marja_row.get("contract", ""))
            cfo_buyer = safe_str(marja_row.get("cfo_buyer", ""))
            if be_sup and contract:
                key = f"{be_sup}{contract}"
                cfo_kv_dict[key] = cfo_buyer

        log.info(f"Создан словарь ЦФО КВ из МАРЖА: {len(cfo_kv_dict)} записей")

        # Check после первого блока (строка end_row_first + 3)
        check_row_1 = end_row_first + 3
        total_sum = sum(safe_float(v.get("sum", 0), 0) for v in filtered_pivot.values())
        ws_calc.cell(row=check_row_1, column=self.START_COL_ACCOUNT + 6, value="Check")
        ws_calc.cell(
            row=check_row_1, column=self.START_COL_ACCOUNT + 7, value=-total_sum
        )

        # ==================== ВТОРОЙ БЛОК (CAPEX) ====================
        current_row = check_row_1 + 2

        # Заголовки CAPEX
        headers_capex = [
            "БЕ + Договор",
            "БЕ поставщика",
            "Договор",
            "ЦФО КВ",
            "Статья КВ",
            "%",
            "Сумма расходов",
            "БЕ покупателя",
        ]

        for col_offset, header in enumerate(headers_capex):
            ws_calc.cell(
                row=current_row,
                column=self.START_COL_ACCOUNT + col_offset,
                value=header,
            )

        # Создаём словарь мэппинга по ключу БЕ+Договор
        # ВАЖНО: map_row[0] может содержать формулу (=AA5&AC5), поэтому
        # создаём ключ из отдельных колонок: map_row[1]=БЕ, map_row[3]=Договор
        mapping_dict = {}
        for map_row in vgo_mapping:
            if not map_row or len(map_row) < 4:
                continue
            be = safe_str(map_row[1])  # БЕ поставщика
            contract = safe_str(map_row[3])  # Договор
            if not be or not contract:
                continue
            key = f"{be}{contract}"
            if key not in mapping_dict:
                mapping_dict[key] = []
            mapping_dict[key].append(map_row)

        log.info(f"Мэппинг ВГО T2: {len(mapping_dict)} ключей")

        # Данные CAPEX
        data_row = current_row + 1
        for key, pv_values in filtered_pivot.items():
            be_supplier = safe_str(pv_values["be_supplier"])
            cfo = safe_str(pv_values["cfo"])  # Договор
            lookup_key = f"{be_supplier}{cfo}"

            # ЦФО КВ берётся из МАРЖА (ЦФО покупателя), а не из мэппинга!
            cfo_kv = cfo_kv_dict.get(lookup_key, "")
            stat_kv = "PI02"  # Константа для CAPEX
            percent_val = 100

            base_sum = safe_float(pv_values.get("sum", 0), 0)
            calculated_sum = -base_sum * percent_val / 100

            # БЕ + Договор
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT, value=f"{be_supplier}{cfo}"
            )
            # БЕ поставщика
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 1, value=be_supplier
            )
            # Договор
            ws_calc.cell(row=data_row, column=self.START_COL_ACCOUNT + 2, value=cfo)
            # ЦФО КВ
            ws_calc.cell(row=data_row, column=self.START_COL_ACCOUNT + 3, value=cfo_kv)
            # Статья КВ
            ws_calc.cell(row=data_row, column=self.START_COL_ACCOUNT + 4, value=stat_kv)
            # %
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 5, value=percent_val
            )
            # Сумма расходов (отрицательная)
            ws_calc.cell(
                row=data_row, column=self.START_COL_ACCOUNT + 6, value=calculated_sum
            )
            # БЕ покупателя
            ws_calc.cell(
                row=data_row,
                column=self.START_COL_ACCOUNT + 7,
                value=pv_values["be_buyer"],
            )

            data_row += 1

        capex_end_row = data_row - 1
        log.info(f"CAPEX блок: {data_row - current_row - 1} записей")

        # Check после CAPEX (через 2 строки) - вычисляем через Python
        # Значение = положительная сумма (противоположная Check после сводной)
        check_row_2 = capex_end_row + 3
        ws_calc.cell(row=check_row_2, column=self.START_COL_ACCOUNT + 6, value="Check")
        ws_calc.cell(
            row=check_row_2, column=self.START_COL_ACCOUNT + 7, value=total_sum
        )

        # ==================== ТРЕТИЙ БЛОК (OPEX) ====================
        current_row = check_row_2 + 2

        # Заголовки OPEX
        headers_opex = [
            "БЕ + Договор",
            "БЕ поставщика",
            "Договор",
            "ЦФО операционное",
            "Статья операционная",
            "%",
            "Сумма расходов",
            "Сумма расходов с накопительным итогом расчетная для формул",
            "БЕ покупателя",
        ]

        for col_offset, header in enumerate(headers_opex):
            ws_calc.cell(
                row=current_row,
                column=self.START_COL_ACCOUNT + col_offset,
                value=header,
            )

        # Данные OPEX
        data_row = current_row + 1
        for key, pv_values in filtered_pivot.items():
            be_supplier = safe_str(pv_values["be_supplier"])
            cfo = safe_str(pv_values["cfo"])  # Договор
            lookup_key = f"{be_supplier}{cfo}"

            if lookup_key in mapping_dict:
                matched_rows = mapping_dict[lookup_key]

                for map_row in matched_rows:
                    # Структура ВГО T2: [0]=БЕ+Договор, [1]=БЕ, [2]=ЦФО опер, [3]=Договор,
                    # [4]=Статья опер, [5]=Расходы, [6]=%
                    cfo_oper = safe_str(map_row[2]) if len(map_row) > 2 else ""
                    stat_oper = safe_str(map_row[4]) if len(map_row) > 4 else ""
                    percent_val = (
                        safe_float(map_row[6], 100) if len(map_row) > 6 else 100
                    )

                    if not cfo_oper:
                        continue

                    base_sum = safe_float(pv_values.get("sum", 0), 0)
                    calculated_sum = base_sum * percent_val / 100

                    # БЕ + Договор
                    ws_calc.cell(
                        row=data_row,
                        column=self.START_COL_ACCOUNT,
                        value=f"{be_supplier}{cfo}",
                    )
                    # БЕ поставщика
                    ws_calc.cell(
                        row=data_row,
                        column=self.START_COL_ACCOUNT + 1,
                        value=be_supplier,
                    )
                    # Договор
                    ws_calc.cell(
                        row=data_row, column=self.START_COL_ACCOUNT + 2, value=cfo
                    )
                    # ЦФО операционное
                    ws_calc.cell(
                        row=data_row, column=self.START_COL_ACCOUNT + 3, value=cfo_oper
                    )
                    # Статья операционная
                    ws_calc.cell(
                        row=data_row, column=self.START_COL_ACCOUNT + 4, value=stat_oper
                    )
                    # %
                    ws_calc.cell(
                        row=data_row,
                        column=self.START_COL_ACCOUNT + 5,
                        value=percent_val,
                    )
                    # Сумма расходов
                    ws_calc.cell(
                        row=data_row,
                        column=self.START_COL_ACCOUNT + 6,
                        value=calculated_sum,
                    )
                    # Сумма расходов с накопительным итогом
                    ws_calc.cell(
                        row=data_row, column=self.START_COL_ACCOUNT + 7, value=base_sum
                    )
                    # БЕ покупателя
                    ws_calc.cell(
                        row=data_row,
                        column=self.START_COL_ACCOUNT + 8,
                        value=pv_values["be_buyer"],
                    )

                    data_row += 1

        log.info(f"OPEX блок: {data_row - current_row - 1} записей")
        log.info(f"Создано 2 блока для ВГО T2: CAPEX и OPEX")


def process_vgo_sheet_full(
    wb_macros,
    ws_calc: Worksheet,
    ws_map: Worksheet,
    ws_marja: Worksheet,
    ws_spr: Worksheet,
    pivot_data: Dict,
    marja_cols: Any,
    start_row_tbl: int,
    exclusions: List[str],
    is_t2: bool = False,
) -> VgoMappingResult:
    """
    Полная обработка листа ВГО (аналог блоков 1.8-1.11 из VBA)

    Args:
        wb_macros: ExcelHandler основной книги
        ws_calc: Лист расчёта
        ws_map: Лист мэппинга
        ws_marja: Лист МАРЖА
        ws_spr: Справочник статей
        pivot_data: Данные из сводной таблицы
        marja_cols: Информация о колонках МАРЖА
        start_row_tbl: Начальная строка таблицы МАРЖА
        exclusions: Список исключаемых статей
        is_t2: True для T2 листов

    Returns:
        VgoMappingResult с результатами обработки
    """
    processor = VgoProcessor(wb_macros, exclusions)

    # 1. Создаём эмуляцию сводной ВГО
    vgo_pivot = processor.create_vgo_pivot_data()

    # 2. Получаем данные мэппинга ВГО
    vgo_mapping = processor.get_vgo_mapping_data(ws_map, is_t2)

    # 3. Находим проблемные записи (Нет в мэппинге)
    arr_err_be = processor.find_err_be(ws_calc, vgo_mapping, is_t2)

    # 4. Собираем данные из МАРЖА для ВГО
    arr_vgo_marja = processor.collect_vgo_marja_data(
        ws_marja, marja_cols, start_row_tbl
    )

    # 5. Добавляем номера договоров к проблемным записям
    arr_err_be = processor.enrich_err_be_with_contracts(arr_err_be, arr_vgo_marja)

    # 6. Поиск в сводной ВГО
    if vgo_pivot is not None:
        arr_rpt = processor.search_in_vgo_pivot(arr_err_be, vgo_pivot)

        # 7. Автозаполнение мэппинга - ОТКЛЮЧЕНО
        # В эталоне нет автоматически добавленных строк в мэппинг
        # processor.autofill_vgo_mapping(ws_map, arr_rpt, ws_spr, is_t2)

        # Получаем данные мэппинга (без автозаполнения)
        vgo_mapping = processor.get_vgo_mapping_data(ws_map, is_t2)

    # 8. Обновляем статусы на листе расчёта
    processor.update_analysis_status(ws_calc, arr_err_be, processor.result.arr_rpt)

    # 9. Создаём блоки мэппинга
    if is_t2:
        # Для T2 листов создаём два блока: CAPEX и OPEX
        # Собираем данные МАРЖА для T2 (с ЦФО КВ = ЦФО покупателя)
        arr_vgo_marja_t2 = processor.collect_vgo_marja_data_t2(
            ws_marja, marja_cols, start_row_tbl
        )
        processor.create_vgo_t2_blocks(
            ws_calc, pivot_data, vgo_mapping, arr_vgo_marja_t2
        )
    else:
        # Для обычных ВГО листов - только один блок OPEX
        processor.create_vgo_second_block(ws_calc, pivot_data, vgo_mapping, is_t2)

    return processor.result
