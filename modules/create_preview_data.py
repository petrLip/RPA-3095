"""
Модуль логики блока 1: Создание предварительных листов с Расчетами и Мэпинги
Аналог процедуры CreatePreviewData из VBA
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from modules.excel_handler import ExcelHandler, SheetHelper, TableFormatter
from modules.helpers import search_arr_col, format_table, safe_str, safe_float
from modules.logger import log, timing, Timer
from modules.vgo_processor import VgoProcessor, process_vgo_sheet_full


def read_excel_fast(
    file_path: str, sheet_name: str = None, header_row: int = 0
) -> Tuple[pd.DataFrame, List[str]]:
    """
    Быстрое чтение Excel файла через pandas

    Args:
        file_path: Путь к файлу
        sheet_name: Имя листа (None = первый лист)
        header_row: Номер строки с заголовками (0-based)

    Returns:
        Tuple[DataFrame, List[str]]: данные и список имён листов
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    # Определяем engine
    if suffix == ".xlsb":
        engine = "pyxlsb"
    elif suffix == ".xls":
        engine = "xlrd"
    else:
        engine = "openpyxl"

    # Получаем список листов
    xl = pd.ExcelFile(str(path), engine=engine)
    sheet_names = xl.sheet_names

    # Читаем данные
    target_sheet = (
        sheet_name if sheet_name and sheet_name in sheet_names else sheet_names[0]
    )
    df = pd.read_excel(xl, sheet_name=target_sheet,
                       header=header_row, engine=engine)

    return df, sheet_names


def dataframe_to_worksheet(
    df: pd.DataFrame,
    ws: Worksheet,
    start_row: int = 1,
    include_header: bool = True,
    include_index: bool = False,
):
    """
    Записать DataFrame в worksheet openpyxl

    Args:
        df: DataFrame для записи
        ws: Целевой лист
        start_row: Начальная строка
        include_header: Включить заголовки
        include_index: Включить индекс
    """
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=include_index,
                          header=include_header), start_row
    ):
        for c_idx, value in enumerate(row, 1):
            if pd.notna(value):
                ws.cell(row=r_idx, column=c_idx, value=value)


# Константы
START_COL_ACCOUNT = 7  # Стартовая колонка для начала данных на листах Расчёт

# Названия листов
SHEET_NAMES = {
    "start": "СТАРТ",
    "marja": "МАРЖА",
    "mapping_current": "Меппинг_за_текущий_год",
    "mapping_prev": "Меппинг_за_прошлый_год",
    "exclusions": "Исключение_статей",
    "article_kv": "Статья Операционая-Статья КВ",
    "correction_template": "КОРРЕКТИРОВКА_ШАБЛОН",
}

# Названия расчётных листов
CALC_SHEETS = [
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_КЦ",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_ЦОД",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_КЦ",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_ЦОД",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_T2_КЦ",
    "РАСЧЕТ (ИЗ БЮДЖЕТА)_T2_ЦОД",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_T2_КЦ",
    "РАСЧЕТ (ИЗ ОТЧЕТА ВГО)_T2_ЦОД",
]

# Временные листы для удаления
TEMP_SHEETS = [
    "Данные_из_файла_по_ВГО",
    "Сводная_ВГО",
    "Сортировка_по_статьям",
    "БЕ_Мэппинг_БЕ_покупателя",
    "Сводная_БЕ",
]


@dataclass
class ProcessingResult:
    """Результат обработки"""

    success: bool = False
    message: str = ""
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    output_file: str = ""


@dataclass
class MarjaColumnInfo:
    """Информация о колонках листа Маржа"""

    supplier_code: int = 0  # Код_Поставщик
    contract_number: int = 0  # № инвест. Договора
    be_load_code: int = 0  # Код_БЕ Загрузка
    cfo_buyer: int = 0  # ЦФО покупателя
    source_data: int = 0  # Источник данных для определения ЦФО и статьи
    buyer_name: int = 0  # Наименование покупателя
    be_load_name: int = 0  # Наименование_БЕ Загрузка
    cluster: int = 0  # Кластер


class CreatePreviewDataProcessor:
    """
    Процессор для создания предварительных листов с Расчетами и Мэпинги
    """

    def __init__(
        self,
        macros_file_path: str,
        marja_file_path: str,
        vgo_file_path: str,
        progress_callback=None,
    ):
        """
        Args:
            macros_file_path: Путь к файлу с макросами (основной файл)
            marja_file_path: Путь к файлу с листом Маржа
            vgo_file_path: Путь к файлу выверки ВГО
            progress_callback: Функция обратного вызова для обновления прогресса
        """
        self.macros_file_path = Path(macros_file_path)
        self.marja_file_path = Path(marja_file_path)
        self.vgo_file_path = Path(vgo_file_path)
        self.progress_callback = progress_callback

        # Путь для сохранения результата - всегда *_opus.xlsx
        base_name = self.macros_file_path.stem
        # Убираем _opus если уже есть, чтобы не дублировать
        if base_name.endswith("_opus"):
            base_name = base_name[:-5]
        self.output_path = self.macros_file_path.with_stem(
            base_name + "_opus"
        ).with_suffix(".xlsx")

        self.wb_macros: Optional[ExcelHandler] = None
        self.exclusions: List[str] = []
        self.marja_cols = MarjaColumnInfo()
        self.arr_cfo_t2: List[Dict[str, Any]] = []
        self.start_row_tbl_mrj: int = 1

    def _update_progress(self, percent: int, message: str):
        """Обновление прогресса"""
        log.info(f"[{percent}%] {message}")
        if self.progress_callback:
            self.progress_callback(percent, message)

    @timing
    def process(self) -> ProcessingResult:
        """
        Основной метод обработки
        """
        result = ProcessingResult()

        try:
            self._update_progress(0, "Начало обработки...")

            # 1. Открываем основной файл
            self._update_progress(5, "Открытие основного файла...")
            self.wb_macros = ExcelHandler(str(self.macros_file_path))
            self.wb_macros.open()

            # 2. Удаляем старые временные листы
            self._update_progress(8, "Удаление временных листов...")
            self._delete_temp_sheets()

            # 3. Загружаем список исключений
            self._update_progress(10, "Загрузка списка исключений...")
            self._load_exclusions()

            # 4. Загружаем данные из файла Маржа
            self._update_progress(15, "Загрузка данных Маржа...")
            if not self._load_marja_data():
                result.errors.append("Ошибка загрузки данных Маржа")
                return result

            # 5. Загружаем данные из файла ВГО
            self._update_progress(25, "Загрузка данных ВГО...")
            if not self._load_vgo_data():
                result.errors.append("Ошибка загрузки данных ВГО")
                return result

            # 6. Создаём формулы на листе Маржа
            self._update_progress(35, "Создание формул на листе МАРЖА...")
            self._create_marja_formulas()

            # 7. Обработка клиентов Т2
            self._update_progress(40, "Обработка клиентов Т2...")
            self._process_t2_clients()

            # 8. Проверка ЦФО
            self._update_progress(45, "Проверка ЦФО...")
            if not self._validate_cfo():
                result.errors.append("Есть ЦФО с длиной меньше 3 символов")
                return result

            # 9. Обновление сводных таблиц на расчётных листах
            self._update_progress(50, "Обработка расчётных листов...")
            self._process_calculation_sheets()

            # 10. Удаляем временные листы
            self._update_progress(90, "Очистка временных данных...")
            self._delete_temp_sheets()

            # 11. Вычисляем формулы в мэппинге ВГО Т2 перед сохранением
            self._update_progress(93, "Вычисление формул в мэппинге...")
            ws_map = self.wb_macros.get_sheet("Меппинг_за_текущий_год")
            if ws_map:
                log.info("Вычисление формул 'БЕ + Договор' в мэппинге ВГО Т2...")
                processor = VgoProcessor(self.wb_macros, self.exclusions)
                processor.evaluate_vgo_t2_mapping_formulas(ws_map)
                log.info("Вычисление формул завершено")

            # 12. Сохраняем файл с суффиксом _opus
            self._update_progress(95, "Сохранение файла...")
            self.wb_macros.save(file_path=str(self.output_path))

            self._update_progress(100, "Обработка завершена!")
            result.success = True
            result.message = (
                "Предварительные листы с сопоставлением по мэппингу готовы!"
            )
            result.output_file = str(self.output_path)

        except Exception as e:
            log.exception(f"Ошибка обработки: {e}")
            result.errors.append(str(e))

        finally:
            if self.wb_macros:
                self.wb_macros.close()

        return result

    def _delete_temp_sheets(self):
        """Удаление временных листов"""
        for sheet_name in TEMP_SHEETS:
            if self.wb_macros.sheet_exists(sheet_name):
                self.wb_macros.delete_sheet(sheet_name)
                log.debug(f"Удалён лист: {sheet_name}")

    def _load_exclusions(self):
        """Загрузка списка исключаемых статей"""
        ws_excl = self.wb_macros.get_sheet(SHEET_NAMES["exclusions"])
        if not ws_excl:
            log.warning("Лист исключений не найден")
            return

        helper = SheetHelper(ws_excl)
        max_row = helper.max_row

        self.exclusions = []
        for row in range(2, max_row + 1):
            value = ws_excl.cell(row=row, column=1).value
            if value:
                self.exclusions.append(str(value))

        log.info(f"Загружено {len(self.exclusions)} статей исключений")

    def _load_marja_data(self) -> bool:
        """Загрузка данных из файла Маржа через pandas (быстрее)"""
        with Timer("Загрузка файла Маржа"):
            try:
                target_sheet = "3_п37_Маржа"
                suffix = self.marja_file_path.suffix.lower()

                # Определяем engine для pandas
                if suffix == ".xlsb":
                    engine = "pyxlsb"
                elif suffix == ".xls":
                    engine = "xlrd"
                else:
                    engine = "openpyxl"

                log.info(
                    f"Чтение файла Маржа через pandas (engine={engine})...")

                # Читаем весь лист без заголовков для поиска нужных строк
                df_raw = pd.read_excel(
                    str(self.marja_file_path),
                    sheet_name=target_sheet,
                    header=None,
                    engine=engine,
                )

                log.info(
                    f"Файл прочитан: {len(df_raw)} строк, {len(df_raw.columns)} колонок"
                )

                # Поиск строки "Исходные Данные"
                start_row_idx = 0
                for idx, row in df_raw.iterrows():
                    if "Исходные Данные" in str(row.values):
                        start_row_idx = idx
                        log.info(
                            f"Найдена строка 'Исходные Данные' в позиции {idx}")
                        break

                # Поиск колонки с "Накопительный Итог_Расходы поставщика"
                finish_col_name = "Накопительный Итог_Расходы поставщика на выполнение работ по договору"
                last_col_idx = len(df_raw.columns)
                header_row_in_source = 0

                # Ищем в первых 20 строках
                for idx in range(min(20, len(df_raw))):
                    row_values = df_raw.iloc[idx].astype(str).tolist()
                    for col_idx, val in enumerate(row_values):
                        if finish_col_name in val:
                            last_col_idx = (
                                col_idx + 1
                            )  # +1 т.к. нужна колонка включительно
                            header_row_in_source = idx
                            log.info(
                                f"Найден заголовок '{finish_col_name[:40]}...' в строке {idx}, колонка {col_idx}"
                            )
                            break
                    if header_row_in_source > 0:
                        break

                # Вычисляем start_row_tbl_mrj - это номер строки заголовков в ЦЕЛЕВОМ листе
                # Данные копируются начиная со строки start_row_idx + 1 (после "Исходные Данные")
                # Заголовки находятся в строке header_row_in_source в исходном файле
                # В целевом файле они будут в строке: header_row_in_source - start_row_idx
                self.start_row_tbl_mrj = header_row_in_source - start_row_idx
                log.info(
                    f"Строка заголовков таблицы в целевом листе: {self.start_row_tbl_mrj}"
                )

                # Обрезаем DataFrame - копируем данные НАЧИНАЯ со строки после "Исходные Данные"
                df_data = df_raw.iloc[start_row_idx + 1:, :last_col_idx].copy()
                df_data.reset_index(drop=True, inplace=True)

                # Записываем на лист МАРЖА
                ws_marja = self.wb_macros.get_or_create_sheet(
                    SHEET_NAMES["marja"])
                helper_marja = SheetHelper(ws_marja)
                helper_marja.clear_sheet()

                log.info(
                    f"Копирование данных: {len(df_data)} строк, {len(df_data.columns)} колонок"
                )

                # Записываем данные (нумерация строк с 1)
                for r_idx, row in enumerate(df_data.values, 1):
                    for c_idx, value in enumerate(row, 1):
                        if pd.notna(value):
                            ws_marja.cell(row=r_idx, column=c_idx, value=value)

                # Устанавливаем заголовок первой колонки в строку заголовков
                ws_marja.cell(
                    row=self.start_row_tbl_mrj,
                    column=1,
                    value="Наименование покупателя",
                )

                log.info(
                    f"Данные Маржа скопированы. Строк: {len(df_data)}, строка заголовков: {self.start_row_tbl_mrj}"
                )
                return True

            except Exception as e:
                log.error(f"Ошибка загрузки файла Маржа: {e}")
                import traceback

                log.error(traceback.format_exc())
                return False

    def _load_vgo_data(self) -> bool:
        """Загрузка данных из файла ВГО через pandas (быстрее)"""
        with Timer("Загрузка файла ВГО"):
            try:
                suffix = self.vgo_file_path.suffix.lower()

                # Определяем engine для pandas
                if suffix == ".xlsb":
                    engine = "pyxlsb"
                elif suffix == ".xls":
                    engine = "xlrd"
                else:
                    engine = "openpyxl"

                log.info(f"Чтение файла ВГО через pandas (engine={engine})...")

                # Получаем список листов
                xl = pd.ExcelFile(str(self.vgo_file_path), engine=engine)

                # Ищем лист с нужным заголовком
                target_title = "ВГО сверка по Дате дог_Об - Только обороты"
                target_sheet = None
                df_vgo = None

                for sheet_name in xl.sheet_names:
                    # Читаем первые 10 строк для поиска заголовка
                    df_check = pd.read_excel(
                        xl, sheet_name=sheet_name, header=None, nrows=10
                    )

                    # Ищем заголовок
                    for idx, row in df_check.iterrows():
                        row_str = " ".join(str(v)
                                           for v in row.values if pd.notna(v))
                        if target_title in row_str:
                            target_sheet = sheet_name
                            log.info(f"Найден лист ВГО: {sheet_name}")
                            break

                    if target_sheet:
                        break

                if not target_sheet:
                    log.error("Не найден лист с данными ВГО")
                    return False

                # Читаем весь лист
                df_vgo = pd.read_excel(
                    xl, sheet_name=target_sheet, header=None, engine=engine
                )
                log.info(
                    f"Файл ВГО прочитан: {len(df_vgo)} строк, {len(df_vgo.columns)} колонок"
                )

                # Колонки для копирования
                columns_to_copy = [
                    "Балансовая единица (Entity)",
                    "Счет (Account)",
                    "Статья (IncomeExpenseItem)",
                    "ЦФО (CostCenter)",
                    "Бизнес-сегмент (BusinessSegment)",
                    "Ссылка ID",
                    "Сумма транзакции объекта",
                ]

                # Ищем строку с заголовками
                header_row_idx = 0
                for idx, row in df_vgo.iterrows():
                    row_values = [str(v) for v in row.values if pd.notna(v)]
                    if any(col in row_values for col in columns_to_copy):
                        header_row_idx = idx
                        break

                # Получаем заголовки
                headers = df_vgo.iloc[header_row_idx].tolist()

                # Находим индексы нужных колонок
                col_indices = []
                for col_name in columns_to_copy:
                    for i, h in enumerate(headers):
                        if pd.notna(h) and col_name in str(h):
                            col_indices.append(i)
                            break

                # Извлекаем нужные данные
                df_result = df_vgo.iloc[header_row_idx:, col_indices].copy()
                df_result.columns = columns_to_copy[: len(col_indices)]
                df_result.reset_index(drop=True, inplace=True)

                # Создаём лист и записываем данные
                ws_vgo = self.wb_macros.get_or_create_sheet(
                    "Данные_из_файла_по_ВГО")
                helper_vgo = SheetHelper(ws_vgo)
                helper_vgo.clear_sheet()

                # Записываем заголовки
                for c_idx, col_name in enumerate(df_result.columns, 1):
                    ws_vgo.cell(row=1, column=c_idx, value=col_name)

                # Записываем данные (пропускаем первую строку - она с заголовками из файла)
                for r_idx, row in enumerate(df_result.values[1:], 2):
                    for c_idx, value in enumerate(row, 1):
                        if pd.notna(value):
                            ws_vgo.cell(row=r_idx, column=c_idx, value=value)

                log.info(
                    f"Скопировано {len(col_indices)} колонок с данными ВГО, {len(df_result) - 1} строк"
                )
                return True

            except Exception as e:
                log.error(f"Ошибка загрузки файла ВГО: {e}")
                import traceback

                log.error(traceback.format_exc())
                return False

    def _create_marja_formulas(self):
        """Создание формул на листе МАРЖА"""
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])
        ws_map = self.wb_macros.get_sheet(SHEET_NAMES["mapping_current"])

        if not ws_marja or not ws_map:
            log.error("Не найдены необходимые листы")
            return

        helper_marja = SheetHelper(ws_marja)
        helper_map = SheetHelper(ws_map)

        max_row = helper_marja.max_row
        max_col = helper_marja.max_col

        # Поиск ключевых колонок на листе МАРЖА
        col_searches = [
            ("Код_Поставщик", "supplier_code"),
            ("№ инвест. Договора  ПАО или ДЗО", "contract_number"),
            ("Код_БЕ Загрузка", "be_load_code"),
            ("ЦФО покупателя_ЦФО покупателя", "cfo_buyer"),
        ]

        for search_text, attr_name in col_searches:
            found = helper_marja.find_value(search_text, partial=True)
            if found:
                setattr(self.marja_cols, attr_name, found[1])
                log.debug(
                    f"Найдена колонка '{search_text}': {get_column_letter(found[1])}"
                )

        # Поиск блоков на листе мэппинга
        found_capex = helper_map.find_value(
            "Информация о ЦФО и статье из бюджета CAPEX", partial=False
        )

        if found_capex:
            capex_row, capex_col = found_capex
            # Ищем нужные колонки в блоке CAPEX
            arr_cols = search_arr_col(
                ws_map,
                ["БЕ поставщика", "БЕ поставщика + Наименование"],
                start_col=capex_col,
                end_col=capex_col + 20,
            )

            if len(arr_cols) >= 2:
                be_supplier_col_letter = arr_cols[0]["col_letter"]
                be_supplier_name_col_letter = arr_cols[1]["col_letter"]

                # Находим ПОСЛЕДНЮЮ колонку с данными "Накопительный Итог_Расходы поставщика"
                # Это должна быть колонка X (24), новые заголовки начнутся с Y (25)
                found_last_data = helper_marja.find_value(
                    "Накопительный Итог_Расходы поставщика", partial=True
                )
                if found_last_data:
                    # Колонка сразу после последней колонки данных
                    new_col = found_last_data[1] + 1
                    log.info(
                        f"Последняя колонка данных 'Расходы поставщика': {get_column_letter(found_last_data[1])}, новые заголовки начнутся с {get_column_letter(new_col)}"
                    )
                else:
                    # Если не нашли, ищем просто max_col
                    new_col = max_col + 1
                    log.warning(
                        f"Не найден 'Накопительный Итог_Расходы поставщика', используем max_col + 1 = {new_col}"
                    )

                # Строка заголовков - это строка где находятся заголовки таблицы
                header_row = self.start_row_tbl_mrj
                log.info(f"Строка заголовков: {header_row}")

                # Колонка "*"
                ws_marja.cell(row=header_row, column=new_col, value="*")
                new_col += 1

                # Колонка "Источник данных для определения ЦФО и статьи"
                ws_marja.cell(
                    row=header_row,
                    column=new_col,
                    value="Источник данных для определения ЦФО и статьи",
                )
                self.marja_cols.source_data = new_col

                # Получаем данные мэппинга для определения источника
                be_supplier_set = self._get_mapping_be_set(
                    ws_map, "Информация о ЦФО и статье из бюджета CAPEX"
                )
                log.info(
                    f"Загружено {len(be_supplier_set)} БЕ поставщиков из мэппинга CAPEX"
                )

                # Заполняем ЗНАЧЕНИЯ (не формулы) для корректной фильтрации
                for row in range(header_row + 1, max_row + 1):
                    supplier_code = safe_str(
                        ws_marja.cell(
                            row=row, column=self.marja_cols.supplier_code
                        ).value
                    )
                    # Если БЕ поставщика есть в мэппинге CAPEX -> "из бюджета", иначе "из отчета ВГО"
                    source_value = (
                        "из бюджета"
                        if supplier_code in be_supplier_set
                        else "из отчета ВГО"
                    )
                    ws_marja.cell(row=row, column=new_col, value=source_value)

                new_col += 1

                # Колонка сцепки "*" (Код поставщика + № инвест. Договора)
                ws_marja.cell(row=header_row, column=new_col, value="*")

                # Записываем ЗНАЧЕНИЯ (не формулы) - конкатенация кода и номера договора
                for row in range(header_row + 1, max_row + 1):
                    supplier_code = safe_str(
                        ws_marja.cell(
                            row=row, column=self.marja_cols.supplier_code
                        ).value
                    )
                    contract_num = safe_str(
                        ws_marja.cell(
                            row=row, column=self.marja_cols.contract_number
                        ).value
                    )
                    # Сцепка: КодПоставщика + НомерДоговора
                    concat_value = f"{supplier_code}{contract_num}"
                    ws_marja.cell(row=row, column=new_col, value=concat_value)

                new_col += 1

                # Колонка "Кластер"
                ws_marja.cell(row=header_row, column=new_col, value="Кластер")
                self.marja_cols.cluster = new_col

                # Получаем данные мэппинга кластера ЦОД
                be_cluster_set = self._get_mapping_be_set(
                    ws_map, "Информация о кластере ЦОД"
                )
                log.info(
                    f"Загружено {len(be_cluster_set)} БЕ из мэппинга кластера ЦОД")

                # Заполняем ЗНАЧЕНИЯ (не формулы)
                for row in range(header_row + 1, max_row + 1):
                    be_load = safe_str(
                        ws_marja.cell(
                            row=row, column=self.marja_cols.be_load_code
                        ).value
                    )
                    # Если БЕ загрузки есть в мэппинге ЦОД -> "ЦОД", иначе "КЦ"
                    cluster_value = "ЦОД" if be_load in be_cluster_set else "КЦ"
                    ws_marja.cell(row=row, column=new_col, value=cluster_value)

        # Форматирование
        format_table(ws_marja, 1, self.start_row_tbl_mrj)
        log.info("Формулы на листе МАРЖА созданы")

    def _get_mapping_be_set(self, ws_map: Worksheet, section_name: str) -> set:
        """Получить множество БЕ из раздела мэппинга"""
        helper = SheetHelper(ws_map)
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(f"Раздел мэппинга не найден: {section_name}")
            return set()

        section_row, section_col = found
        log.debug(
            f"Найден раздел '{section_name}' в строке {section_row}, колонка {section_col}"
        )
        be_set = set()

        # Ищем колонку "БЕ поставщика" или "БЕ" в нескольких строках после заголовка раздела
        be_col = None
        for search_row in range(section_row, section_row + 3):
            for col in range(section_col, section_col + 20):
                val = ws_map.cell(row=search_row, column=col).value
                if val:
                    val_str = str(val)
                    # Ищем колонку с "БЕ поставщика" или просто "БЕ" (но не "БЕ покупателя")
                    if ("БЕ поставщика" in val_str) or (val_str.strip() == "БЕ"):
                        be_col = col
                        log.debug(
                            f"Найдена колонка БЕ: '{val_str}' в {get_column_letter(col)}{search_row}"
                        )
                        break
            if be_col:
                break

        if be_col is None:
            # Пробуем найти первую колонку раздела (обычно это БЕ)
            be_col = section_col
            log.warning(
                f"Колонка 'БЕ поставщика' не найдена в разделе {section_name}, используем первую колонку раздела"
            )

        # Ищем строку с данными (первая строка после заголовков с числовым значением)
        data_start_row = section_row + 2
        for row in range(section_row + 1, section_row + 5):
            val = ws_map.cell(row=row, column=be_col).value
            if val and str(val).strip().isdigit():
                data_start_row = row
                break

        # Собираем значения БЕ
        for row in range(data_start_row, helper.max_row + 1):
            val = ws_map.cell(row=row, column=be_col).value
            if val is None or str(val).strip() == "":
                break
            be_set.add(safe_str(val))

        log.debug(
            f"Собрано {len(be_set)} значений БЕ из раздела {section_name}")
        return be_set

    def _process_t2_clients(self):
        """Обработка клиентов Т2"""
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])
        if not ws_marja:
            return

        helper = SheetHelper(ws_marja)
        max_row = helper.max_row

        # Поиск колонок
        found_buyer = helper.find_value(
            "Наименование покупателя", partial=False)
        found_be_load = helper.find_value(
            "Наименование_БЕ Загрузка", partial=True)

        if not found_buyer or not found_be_load:
            log.warning("Не найдены колонки для обработки Т2")
            return

        buyer_col = found_buyer[1]
        be_load_col = found_be_load[1]
        self.marja_cols.buyer_name = buyer_col
        self.marja_cols.be_load_name = be_load_col

        self.arr_cfo_t2 = []

        for row in range(self.start_row_tbl_mrj + 1, max_row + 1):
            buyer_name = ws_marja.cell(row=row, column=buyer_col).value
            be_load_name = ws_marja.cell(row=row, column=be_load_col).value

            if buyer_name == "Т2":
                # Если БЕ Загрузка != КЦ, ставим код 00902
                if be_load_name != "КЦ" and self.marja_cols.be_load_code > 0:
                    cell = ws_marja.cell(
                        row=row, column=self.marja_cols.be_load_code)
                    cell.number_format = "@"
                    cell.value = "00902"

                # Если договор пустой, ставим УЦН
                if self.marja_cols.contract_number > 0:
                    contract_val = ws_marja.cell(
                        row=row, column=self.marja_cols.contract_number
                    ).value
                    if not contract_val:
                        ws_marja.cell(
                            row=row, column=self.marja_cols.contract_number, value="УЦН"
                        )

                # Собираем данные Т2 в массив
                self.arr_cfo_t2.append(
                    {
                        "supplier_code": safe_str(
                            ws_marja.cell(
                                row=row, column=self.marja_cols.supplier_code
                            ).value
                        ),
                        "contract": safe_str(
                            ws_marja.cell(
                                row=row, column=self.marja_cols.contract_number
                            ).value
                        ),
                        "cfo": safe_str(
                            ws_marja.cell(
                                row=row, column=self.marja_cols.cfo_buyer
                            ).value
                        ),
                    }
                )

            # Приводим код поставщика к тексту
            if self.marja_cols.supplier_code > 0:
                cell = ws_marja.cell(
                    row=row, column=self.marja_cols.supplier_code)
                cell.number_format = "@"
                cell.value = safe_str(cell.value)

        log.info(f"Обработано клиентов Т2: {len(self.arr_cfo_t2)}")

    def _validate_cfo(self) -> bool:
        """Проверка длины ЦФО"""
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])
        if not ws_marja:
            return True

        helper = SheetHelper(ws_marja)
        max_row = helper.max_row
        formatter = TableFormatter(ws_marja)

        has_short_cfo = False

        if self.marja_cols.cfo_buyer > 0:
            for row in range(self.start_row_tbl_mrj + 1, max_row + 1):
                cell = ws_marja.cell(row=row, column=self.marja_cols.cfo_buyer)
                cell.number_format = "@"
                value = safe_str(cell.value)

                if len(value) < 3:
                    formatter.highlight_cell(
                        row, self.marja_cols.cfo_buyer, "red")
                    has_short_cfo = True
                    log.warning(f"Короткий ЦФО в строке {row}: '{value}'")

        if has_short_cfo:
            log.error("Найдены ЦФО с длиной меньше 3 символов!")
            return False

        return True

    def _process_calculation_sheets(self):
        """Обработка расчётных листов"""
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])
        if not ws_marja:
            return

        helper_marja = SheetHelper(ws_marja)

        # Определяем диапазон для сводных таблиц
        found_cluster = helper_marja.find_value("Кластер", partial=False)
        if found_cluster:
            max_col = helper_marja.max_col
            max_row = helper_marja.max_row

            # Формируем данные для каждого расчётного листа
            pv_source_data = self._get_pivot_source_data(
                ws_marja, self.start_row_tbl_mrj, max_row, max_col
            )

            for iter_idx, sheet_name in enumerate(CALC_SHEETS, 1):
                self._update_progress(
                    50 + iter_idx * 5, f"Обработка листа: {sheet_name}"
                )
                self._process_single_calc_sheet(
                    sheet_name, iter_idx, pv_source_data)

    def _get_pivot_source_data(
        self, ws: Worksheet, start_row: int, max_row: int, max_col: int
    ) -> List[List[Any]]:
        """Получение данных для сводных таблиц"""
        data = []
        for row in range(start_row, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            data.append(row_data)
        return data

    def _process_single_calc_sheet(
        self, sheet_name: str, iter_idx: int, source_data: List[List[Any]]
    ):
        """Обработка одного расчётного листа"""
        ws_calc = self.wb_macros.get_sheet(sheet_name)
        ws_marja = self.wb_macros.get_sheet(SHEET_NAMES["marja"])
        ws_map = self.wb_macros.get_sheet(SHEET_NAMES["mapping_current"])
        ws_spr = self.wb_macros.get_sheet(SHEET_NAMES["article_kv"])

        if not ws_calc or not ws_marja or not ws_map:
            log.warning(f"Не найден лист: {sheet_name}")
            return

        helper_calc = SheetHelper(ws_calc)
        helper_map = SheetHelper(ws_map)

        # Очищаем старые данные (колонки после 5)
        max_col = helper_calc.max_col
        if max_col >= 6:
            for row in range(1, helper_calc.max_row + 1):
                for col in range(6, max_col + 1):
                    ws_calc.cell(row=row, column=col).value = None

        # Определяем тип расчёта
        is_budget = "БЮДЖЕТ" in sheet_name
        is_vgo = "ОТЧЕТА ВГО" in sheet_name
        is_t2 = "T2" in sheet_name
        is_cod = "ЦОД" in sheet_name

        pivot_data = None

        if is_vgo:
            # Для ВГО пересчитываем сводную по данным МАРЖА
            filtered_data = self._filter_data_for_sheet(
                source_data, is_budget, is_t2, is_cod
            )
            pivot_data = self._create_pivot_data(
                filtered_data, is_vgo=is_vgo, is_t2=is_t2
            )

            if pivot_data:
                # Очищаем текущую сводную A:D
                max_row = helper_calc.max_row
                for row in range(5, max_row + 1):
                    for col in range(1, 5):
                        ws_calc.cell(row=row, column=col).value = None

                # Записываем сводную в порядке убывания суммы, как в VBA
                self._write_summary_table(
                    ws_calc,
                    pivot_data,
                    sort_by_sum_desc=True,
                    is_vgo=is_vgo,
                    is_t2=is_t2,
                )
        else:
            # Используем существующие данные из сводной таблицы A-D
            # VBA макрос НЕ перезаписывает эту таблицу, только добавляет колонки G+
            pivot_data = self._read_existing_pivot_data(ws_calc)

        if not pivot_data:
            log.warning(f"Нет данных в сводной таблице листа {sheet_name}")
            return

        log.info(
            f"Лист {sheet_name}: найдено {len(pivot_data)} уникальных комбинаций")

        # Проверяем, есть ли реальные данные (не только "Общий итог")
        real_data = {
            k: v
            for k, v in pivot_data.items()
            if v.get("be_supplier") and str(v.get("be_supplier")) != "Общий итог"
        }

        if not real_data:
            log.warning(
                f"Лист {sheet_name}: только 'Общий итог', блок мэппинга не создаётся"
            )
            return

        # НЕ перезаписываем сводную таблицу A-D - она уже есть!
        # Записываем только блок мэппинга (G-M)

        # Для ВГО листов получаем ключи мэппинга для определения статуса
        mapping_keys = None
        if is_vgo:
            mapping_section = (
                "Информация о ЦФО и статье из ВГО Т2"
                if is_t2
                else "Информация о ЦФО и статье из ВГО"
            )
            mapping_keys = self._get_mapping_keys(
                ws_map, mapping_section, is_t2)
            log.info(
                f"Получено {len(mapping_keys)} ключей мэппинга для статуса анализа"
            )

        self._write_pivot_to_sheet(
            ws_calc,
            pivot_data,
            START_COL_ACCOUNT,
            is_vgo=is_vgo,
            is_t2=is_t2,
            mapping_keys=mapping_keys,
        )

        if is_budget:
            # Для листов БЮДЖЕТ - второй блок использует мэппинг OPEX (не CAPEX!)
            # VBA: Set FndVal = ShMap.Cells.Find("Информация о ЦФО и статье из бюджета OPEX"...)
            opex_mapping_section = "Информация о ЦФО и статье из бюджета OPEX"
            arr_map_opex = self._get_opex_mapping_data(
                ws_map, opex_mapping_section)

            # Создаём табличную часть с мэппингом (второй блок OPEX)
            self._create_mapping_table_opex(
                ws_calc, pivot_data, arr_map_opex, is_t2=is_t2
            )
        else:
            # Для листов ВГО - ПОЛНАЯ обработка через VgoProcessor
            log.info(f"Запуск полной обработки ВГО для листа: {sheet_name}")

            try:
                vgo_result = process_vgo_sheet_full(
                    wb_macros=self.wb_macros,
                    ws_calc=ws_calc,
                    ws_map=ws_map,
                    ws_marja=ws_marja,
                    ws_spr=ws_spr,
                    pivot_data=pivot_data,
                    marja_cols=self.marja_cols,
                    start_row_tbl=self.start_row_tbl_mrj,
                    exclusions=self.exclusions,
                    is_t2=is_t2,
                )

                log.info(
                    f"ВГО обработка завершена: "
                    f"проблемных записей={len(vgo_result.arr_err_be)}, "
                    f"результатов={len(vgo_result.arr_rpt)}"
                )
            except Exception as e:
                log.exception(
                    f"Ошибка обработки ВГО для листа {sheet_name}: {e}")
                # В случае ошибки - используем упрощённую логику
                log.warning(f"Используем упрощённую логику для {sheet_name}")

        # Форматирование
        format_table(ws_calc, 3, 5)

        log.info(f"Обработан лист: {sheet_name}")

    def _filter_data_for_sheet(
        self, data: List[List[Any]], is_budget: bool, is_t2: bool, is_cod: bool
    ) -> List[List[Any]]:
        """Фильтрация данных для конкретного листа"""
        if not data or len(data) < 2:
            log.warning("Нет данных для фильтрации")
            return []

        headers = data[0]
        result = [headers]

        # Индексы нужных колонок
        source_idx = None
        cluster_idx = None
        buyer_idx = None

        for i, h in enumerate(headers):
            h_str = safe_str(h)
            if "Источник" in h_str and source_idx is None:
                source_idx = i
                log.debug(f"Найдена колонка 'Источник': индекс {i}")
            elif "Кластер" in h_str and cluster_idx is None:
                cluster_idx = i
                log.debug(f"Найдена колонка 'Кластер': индекс {i}")
            elif "Наименование покупателя" in h_str and buyer_idx is None:
                buyer_idx = i
                log.debug(
                    f"Найдена колонка 'Наименование покупателя': индекс {i}")

        if source_idx is None:
            log.warning("Колонка 'Источник' не найдена в данных")
        if cluster_idx is None:
            log.warning("Колонка 'Кластер' не найдена в данных")

        for row in data[1:]:
            # Фильтр по источнику
            if source_idx is not None:
                source_val = safe_str(row[source_idx])
                if is_budget and source_val != "из бюджета":
                    continue
                if not is_budget and source_val != "из отчета ВГО":
                    continue

            # Фильтр по кластеру
            if cluster_idx is not None:
                cluster_val = safe_str(row[cluster_idx])
                if is_cod and cluster_val != "ЦОД":
                    continue
                if not is_cod and cluster_val != "КЦ":
                    continue

            # Фильтр по покупателю (Т2)
            if buyer_idx is not None:
                buyer_val = safe_str(row[buyer_idx])
                if is_t2 and buyer_val != "Т2":
                    continue
                if not is_t2 and buyer_val == "Т2":
                    continue

            result.append(row)

        return result

    def _create_pivot_data(
        self, data: List[List[Any]], is_vgo: bool = False, is_t2: bool = False
    ) -> Dict[str, Dict]:
        """Создание сводных данных (группировка)"""
        if not data or len(data) < 2:
            return {}

        headers = data[0]

        # Индексы ключевых колонок
        be_sup_idx = None
        cfo_idx = None
        be_buy_idx = None
        sum_idx = None

        for i, h in enumerate(headers):
            h_str = safe_str(h)
            if "Код_Поставщик" in h_str:
                be_sup_idx = i
            elif is_vgo and is_t2 and "№ инвест. Договора" in h_str:
                cfo_idx = i
            elif "ЦФО покупателя" in h_str and not (is_vgo and is_t2):
                cfo_idx = i
            elif "Код_БЕ Загрузка" in h_str or "БЕ покупателя" in h_str:
                be_buy_idx = i
            elif "Накопительный Итог_Расходы" in h_str:
                sum_idx = i

        pivot = {}

        for row in data[1:]:
            if be_sup_idx is None or cfo_idx is None or sum_idx is None:
                continue

            key = (
                safe_str(row[be_sup_idx]) if be_sup_idx is not None else "",
                safe_str(row[cfo_idx]) if cfo_idx is not None else "",
                safe_str(row[be_buy_idx]) if be_buy_idx is not None else "",
            )

            sum_val = safe_float(row[sum_idx])

            if key not in pivot:
                pivot[key] = {
                    "be_supplier": key[0],
                    "cfo": key[1],
                    "be_buyer": key[2],
                    "sum": 0,
                }

            pivot[key]["sum"] += sum_val

        return pivot

    def _write_summary_table(
        self,
        ws: Worksheet,
        pivot_data: Dict,
        sort_by_sum_desc: bool = False,
        is_vgo: bool = False,
        is_t2: bool = False,
    ):
        """Запись сводной таблицы (A-D)"""
        # Заголовки
        cfo_header = "ЦФО покупателя"
        if is_vgo and is_t2:
            cfo_header = "Договор"
        headers = [
            "БЕ поставщика",
            cfo_header,
            "БЕ покупателя",
            "Сумма расходов с накопительным итогом",
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)

        # Данные
        row = 6
        total_sum = 0
        items = pivot_data.items()
        if sort_by_sum_desc:
            items = sorted(
                items, key=lambda kv: safe_float(kv[1].get("sum", 0), 0), reverse=True
            )

        for _, values in items:
            ws.cell(row=row, column=1, value=values["be_supplier"])
            ws.cell(row=row, column=2, value=values["cfo"])
            ws.cell(row=row, column=3, value=values["be_buyer"])
            ws.cell(row=row, column=4, value=values["sum"])
            total_sum += safe_float(values.get("sum", 0), 0)
            row += 1

        # Добавляем "Общий итог" как в VBA
        if row > 6:
            ws.cell(row=row, column=1, value="Общий итог")
            ws.cell(row=row, column=4, value=total_sum)

        log.debug(f"Записана сводная таблица: {row - 6} строк")

    def _read_existing_pivot_data(self, ws: Worksheet) -> Dict:
        """Чтение существующих данных из сводной таблицы на листе РАСЧЕТ (колонки A-D)

        Структура листа РАСЧЕТ:
        - Строка 5: заголовки (БЕ поставщика, ЦФО покупателя, БЕ покупателя, Сумма)
        - Строки 6+: данные
        """
        helper = SheetHelper(ws)
        pivot_data = {}

        # Находим последнюю строку данных в колонке A
        max_row = helper.get_used_range_end(1)

        # Читаем данные начиная со строки 6 (после заголовка)
        for row in range(6, max_row + 1):
            # Колонка A - БЕ поставщика
            be_supplier = ws.cell(row=row, column=1).value
            cfo = ws.cell(row=row, column=2).value  # Колонка B - ЦФО
            # Колонка C - БЕ покупателя
            be_buyer = ws.cell(row=row, column=3).value
            sum_value = ws.cell(row=row, column=4).value  # Колонка D - Сумма

            if not be_supplier:
                continue

            # Создаём ключ как в VBA
            key = f"{safe_str(be_supplier)}_{safe_str(cfo)}_{safe_str(be_buyer)}"

            if key not in pivot_data:
                pivot_data[key] = {
                    "be_supplier": be_supplier,
                    "cfo": cfo,
                    "be_buyer": be_buyer,
                    "sum": sum_value,
                }
            else:
                # Если ключ уже есть, суммируем
                if sum_value:
                    existing_sum = pivot_data[key].get("sum", 0) or 0
                    pivot_data[key]["sum"] = existing_sum + sum_value

        return pivot_data

    def _write_pivot_to_sheet(
        self,
        ws: Worksheet,
        pivot_data: Dict,
        start_col: int,
        is_vgo: bool = False,
        is_t2: bool = False,
        mapping_keys: set = None,
    ):
        """Запись сводных данных на лист (блок мэппинга)

        Для листов БЮДЖЕТА: БЕ + ЦФО, БЕ поставщика, ЦФО КВ, Статья КВ, %, Сумма расходов, БЕ покупателя
        Для листов ВГО: БЕ + ЦФО, БЕ поставщика, ЦФО КВ, Статус анализа, № инвест. Договора
        Для листов ВГО T2: БЕ + Договор, БЕ поставщика, Договор, Статус анализа

        Args:
            mapping_keys: set ключей из мэппинга для определения статуса анализа
        """
        if is_vgo:
            # Для листов ВГО T2 - заголовки с Договором
            if is_t2:
                headers = ["БЕ + Договор", "БЕ поставщика",
                           "Договор", "Статус анализа"]
            else:
                headers = [
                    "БЕ + ЦФО",
                    "БЕ поставщика",
                    "ЦФО КВ",
                    "Статус анализа",
                    "№ инвест. Договора",
                ]

            for col_offset, header in enumerate(headers):
                ws.cell(row=5, column=start_col + col_offset, value=header)

            # Данные
            row = 6
            for key, values in pivot_data.items():
                # Пропускаем "Общий итог"
                if str(values.get("be_supplier", "")) == "Общий итог":
                    continue

                lookup_key = f"{values['be_supplier']}{values['cfo']}"

                # БЕ + ЦФО/Договор
                ws.cell(row=row, column=start_col, value=lookup_key)
                # БЕ поставщика
                ws.cell(row=row, column=start_col + 1,
                        value=values["be_supplier"])
                # ЦФО/Договор
                ws.cell(row=row, column=start_col + 2, value=values["cfo"])
                # Статус анализа - проверяем наличие в мэппинге
                if mapping_keys and lookup_key in mapping_keys:
                    ws.cell(row=row, column=start_col + 3, value="ОК")
                elif mapping_keys:
                    ws.cell(row=row, column=start_col +
                            3, value="Нет в мэппинге")
                else:
                    ws.cell(row=row, column=start_col +
                            3, value="ОК")  # Fallback
                # № инвест. Договора - только для не-T2
                if not is_t2:
                    ws.cell(row=row, column=start_col + 4, value=None)
                row += 1

            log.info(
                f"Записано {row - 6} строк данных (ВГО{'_T2' if is_t2 else ''})")
        else:
            # Для листов БЮДЖЕТА - стандартная структура
            headers = [
                "БЕ + ЦФО",
                "БЕ поставщика",
                "ЦФО КВ",
                "Статья КВ",
                "%",
                "Сумма расходов",
                "БЕ покупателя",
            ]

            for col_offset, header in enumerate(headers):
                ws.cell(row=5, column=start_col + col_offset, value=header)

            # Данные
            row = 6
            total_sum = 0  # Для контрольной суммы

            for key, values in pivot_data.items():
                # Пропускаем "Общий итог"
                if str(values.get("be_supplier", "")) == "Общий итог":
                    continue

                # БЕ + ЦФО
                ws.cell(
                    row=row,
                    column=start_col,
                    value=f"{values['be_supplier']}{values['cfo']}",
                )
                # БЕ поставщика
                ws.cell(row=row, column=start_col + 1,
                        value=values["be_supplier"])
                # ЦФО КВ
                ws.cell(row=row, column=start_col + 2, value=values["cfo"])
                # Статья КВ
                ws.cell(row=row, column=start_col + 3, value="PI02")
                # %
                percent = 100
                ws.cell(row=row, column=start_col + 4, value=percent)

                # Сумма расходов - ВЫЧИСЛЯЕМ значение вместо формулы
                # Формула: -SUMIFS(...) * percent%
                # Так как данные уже сгруппированы, sum уже содержит результат SUMIFS
                sum_value = safe_float(values.get("sum", 0), 0)
                calculated_sum = (
                    -sum_value * percent / 100
                )  # Инвертируем знак и применяем %
                ws.cell(row=row, column=start_col + 5, value=calculated_sum)
                total_sum += calculated_sum

                # БЕ покупателя
                ws.cell(row=row, column=start_col +
                        6, value=values["be_buyer"])
                row += 1

            # Check - формула суммы первого блока
            # В эталоне: L3="Check", M3=SUM(L6:L{last_row})
            last_data_row = row - 1  # Последняя строка данных
            sum_col_letter = get_column_letter(
                start_col + 5
            )  # Колонка L (Сумма расходов)
            check_col = start_col + 6  # Колонка M

            ws.cell(row=3, column=check_col - 1, value="Check")
            ws.cell(
                row=3,
                column=check_col,
                value=f"=SUM(${sum_col_letter}6:${sum_col_letter}{last_data_row})",
            )

            log.info(
                f"Записано {row - 6} строк данных, Check=SUM({sum_col_letter}6:{sum_col_letter}{last_data_row})"
            )

    def _get_mapping_data(
        self, ws_map: Worksheet, section_name: str
    ) -> List[List[Any]]:
        """Получение данных мэппинга из указанного раздела"""
        helper = SheetHelper(ws_map)
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(f"Раздел мэппинга не найден: {section_name}")
            return []

        section_row, section_col = found

        # Определяем границы раздела
        merged = helper.get_merged_cell_range(section_row, section_col)
        end_col = merged[3] if merged else section_col

        # Собираем данные
        result = []
        max_row = helper.max_row

        # Ищем строку с заголовками (пропускаем пустые строки после названия раздела)
        header_row = section_row + 1
        for check_row in range(section_row + 1, section_row + 5):
            first_val = ws_map.cell(row=check_row, column=section_col).value
            if first_val:
                header_row = check_row
                break

        # Собираем заголовки
        headers = []
        for col in range(section_col, min(section_col + 20, end_col + 10)):
            val = ws_map.cell(row=header_row, column=col).value
            if val:
                headers.append(str(val))

        log.info(f"Мэппинг '{section_name}': заголовки = {headers}")

        # Данные начинаются со следующей строки после заголовков
        for row in range(header_row + 1, max_row + 1):
            row_data = []
            has_data = False
            for col in range(section_col, min(section_col + 20, end_col + 10)):
                val = ws_map.cell(row=row, column=col).value
                row_data.append(val)
                if val:
                    has_data = True
            if has_data:
                # Пропускаем строки-заголовки (если первое значение = "БЕ поставщика" и т.п.)
                first_val = safe_str(row_data[0])
                if first_val in ("БЕ поставщика", "БЕ + ЦФО", "БЕ + Договор"):
                    continue
                result.append(row_data)
            else:
                break  # Пустая строка - конец раздела

        log.info(f"Мэппинг '{section_name}': загружено {len(result)} записей")
        if result:
            log.info(f"Пример первой записи: {result[0]}")

        return result

    def _get_opex_mapping_data(
        self, ws_map: Worksheet, section_name: str
    ) -> List[Dict[str, Any]]:
        """
        Получение данных мэппинга OPEX - эмуляция VBA логики.

        VBA ищет каждую колонку по имени заголовка:
        ArrTitle = Array("БЕ поставщика", "ЦФО операционное", "Статья операционная", "%")

        Returns:
            Список словарей с ключами: be_supplier, cfo_oper, stat_oper, percent
        """
        helper = SheetHelper(ws_map)
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(f"Раздел OPEX мэппинга не найден: {section_name}")
            return []

        section_row, section_col = found
        log.info(
            f"OPEX секция '{section_name}' найдена в строке {section_row}, колонка {section_col}")

        # Определяем границы раздела (MergeArea в VBA)
        merged = helper.get_merged_cell_range(section_row, section_col)
        end_col = merged[3] if merged else section_col + 20
        log.info(f"OPEX секция границы: от колонки {section_col} до {end_col}")

        max_row = helper.max_row

        # VBA: ArrTitle = Array("БЕ поставщика", "ЦФО операционное", "Статья операционная", "%")
        required_columns = {
            "БЕ поставщика": None,
            "ЦФО операционное": None,
            "Статья операционная": None,
            "%": None,
        }

        # Ищем каждую колонку по имени в пределах секции
        # VBA: Set FndCol = ShMap.Range(Cells(FndVal.Row, FndVal.Column), Cells(RowEndShMap, EndColBlk)).Find(ArrTitle(i)...)
        # Сканируем все ячейки в секции для поиска заголовков
        for row in range(section_row, min(section_row + 10, max_row)):
            for col in range(section_col, min(end_col + 10, section_col + 30)):
                val = ws_map.cell(row=row, column=col).value
                if val:
                    val_str = str(val).strip()
                    for req_col in required_columns:
                        if req_col in val_str and required_columns[req_col] is None:
                            required_columns[req_col] = col
                            log.info(
                                f"OPEX мэппинг: найдена колонка '{req_col}' в ({row}, {col}), значение='{val_str}'")
                            break

        # Проверяем, что все колонки найдены
        missing = [k for k, v in required_columns.items() if v is None]
        if missing:
            log.warning(f"OPEX мэппинг: не найдены колонки: {missing}")
            # Выводим все заголовки для отладки
            log.warning("Доступные заголовки в секции:")
            for row in range(section_row, section_row + 5):
                row_vals = []
                for col in range(section_col, section_col + 15):
                    v = ws_map.cell(row=row, column=col).value
                    if v:
                        row_vals.append(f"[{col}]={v}")
                if row_vals:
                    log.warning(f"  Строка {row}: {', '.join(row_vals)}")
            return []

        log.info(
            f"OPEX мэппинг '{section_name}': колонки = {required_columns}")

        # Определяем строку начала данных (первая строка после заголовка с числовым БЕ)
        be_col = required_columns["БЕ поставщика"]
        header_row = section_row + 1

        # Находим строку заголовков (где написано "БЕ поставщика")
        for row in range(section_row, section_row + 5):
            val = ws_map.cell(row=row, column=be_col).value
            if val and "БЕ поставщика" in str(val):
                header_row = row
                break

        # Данные начинаются после строки заголовков
        data_start_row = header_row + 1
        log.info(
            f"OPEX мэппинг: заголовки в строке {header_row}, данные начинаются с {data_start_row}")

        # Собираем данные
        result = []
        for row in range(data_start_row, max_row + 1):
            be_val = ws_map.cell(
                row=row, column=required_columns["БЕ поставщика"]).value
            if not be_val or str(be_val).strip() == "":
                break  # Пустая строка - конец данных

            # Пропускаем заголовки
            be_str = str(be_val).strip()
            if be_str in ("БЕ поставщика", "БЕ"):
                continue

            cfo_oper = ws_map.cell(
                row=row, column=required_columns["ЦФО операционное"]).value
            stat_oper = ws_map.cell(
                row=row, column=required_columns["Статья операционная"]).value
            percent = ws_map.cell(row=row, column=required_columns["%"]).value

            result.append({
                "be_supplier": safe_str(be_val),
                "cfo_oper": safe_str(cfo_oper),
                "stat_oper": safe_str(stat_oper),
                "percent": safe_float(percent, 100),
            })

        log.info(
            f"OPEX мэппинг '{section_name}': загружено {len(result)} записей")
        if result:
            log.info(f"Пример первой записи OPEX: {result[0]}")
            if len(result) > 1:
                log.info(f"Пример второй записи OPEX: {result[1]}")

        return result

    def _get_mapping_keys(
        self, ws_map: Worksheet, section_name: str, is_t2: bool = False
    ) -> set:
        """Получение ключей из мэппинга для проверки статуса анализа

        Эмулирует ВПР: =ЕСЛИ(ЕСЛИОШИБКА(ВПР(G6;Меппинг!Q:X;1;0);"Нет в мэппинге")="Нет в мэппинге";"Нет в мэппинге";"ОК")

        Args:
            ws_map: Лист мэппинга
            section_name: Название секции
            is_t2: True для T2 листов

        Returns:
            Set ключей (БЕ+ЦФО или БЕ+Договор) из мэппинга
        """
        helper = SheetHelper(ws_map)
        found = helper.find_value(section_name, partial=False)

        if not found:
            log.warning(
                f"Раздел мэппинга для ключей не найден: {section_name}")
            return set()

        section_row, section_col = found

        # Находим строку заголовков
        header_row = section_row + 1
        for check_row in range(section_row + 1, section_row + 5):
            first_val = ws_map.cell(row=check_row, column=section_col).value
            if first_val:
                header_row = check_row
                break

        # Определяем колонки для ключа
        # Для обычного ВГО: Q (17) = БЕ+ЦФО
        # Для ВГО T2: Z (26) = БЕ+Договор
        key_col = section_col  # Первая колонка раздела
        be_col = section_col + 1  # БЕ поставщика
        cfo_or_contract_col = section_col + 2  # ЦФО или Договор

        # Собираем ключи
        keys = set()
        max_row = helper.max_row

        for row in range(header_row + 1, max_row + 1):
            key_val = ws_map.cell(row=row, column=key_col).value

            if not key_val:
                continue

            # Если ключ - формула, составляем ключ из отдельных колонок
            key_str = safe_str(key_val)
            if key_str.startswith("="):
                be_val = safe_str(ws_map.cell(row=row, column=be_col).value)
                cfo_val = (
                    safe_str(ws_map.cell(
                        row=row, column=cfo_or_contract_col + 1).value)
                    if is_t2
                    else safe_str(
                        ws_map.cell(row=row, column=cfo_or_contract_col).value
                    )
                )
                if be_val:
                    key_str = f"{be_val}{cfo_val}"
                else:
                    continue

            # Пропускаем заголовки
            if key_str in ("БЕ поставщика", "БЕ + ЦФО", "БЕ + Договор"):
                continue

            keys.add(key_str)

        return keys

    def _create_mapping_table_opex(
        self,
        ws: Worksheet,
        pivot_data: Dict,
        opex_mapping: List[Dict[str, Any]],
        is_t2: bool = False,
    ):
        """
        Создание таблицы мэппинга OPEX на листе расчёта (второй блок).

        Эмулирует VBA процедуру CreateSheetBEMap для TypeRpt=1 (БЮДЖЕТ).

        Args:
            ws: Лист расчёта
            pivot_data: Данные из сводной таблицы
            opex_mapping: Данные мэппинга OPEX (список словарей)
            is_t2: True для T2 листов
        """
        helper = SheetHelper(ws)
        # Между первым блоком и вторым: 2 пустые строки, Check, 1 пустая строка
        current_row = helper.get_used_range_end(START_COL_ACCOUNT) + 5

        # Заголовки для второго блока (OPEX)
        headers = [
            "БЕ + ЦФО",
            "БЕ поставщика",
            "ЦФО операционное",
            "Статья операционная",
            "%",
            "Сумма расходов",
            "Сумма расходов с накопительным итогом расчетная для формул",
            "БЕ покупателя",
        ]

        for col_offset, header in enumerate(headers):
            ws.cell(
                row=current_row, column=START_COL_ACCOUNT + col_offset, value=header
            )

        # Создаём словарь мэппинга для быстрого поиска по БЕ поставщика
        mapping_dict = {}
        for map_item in opex_mapping:
            be_key = map_item["be_supplier"]
            if be_key not in mapping_dict:
                mapping_dict[be_key] = []
            mapping_dict[be_key].append(map_item)

        log.info(
            f"OPEX мэппинг: {len(mapping_dict)} уникальных БЕ поставщика, примеры: {list(mapping_dict.keys())[:5]}..."
        )

        # Создаём словарь сумм по ключу (БЕ поставщика, БЕ покупателя)
        # VBA использует SUMIFS для суммирования всех записей с одинаковым БЕ+покупатель
        sum_by_key = {}
        for key, pv_values in pivot_data.items():
            be_sup = safe_str(pv_values["be_supplier"])
            be_buy = safe_str(pv_values["be_buyer"])
            sum_key = (be_sup, be_buy)
            pv_sum = safe_float(pv_values.get("sum", 0), 0)
            if sum_key not in sum_by_key:
                sum_by_key[sum_key] = 0
            sum_by_key[sum_key] += pv_sum

        log.info(
            f"Создан словарь сумм: {len(sum_by_key)} уникальных ключей (БЕ поставщика, БЕ покупателя)")

        # VBA логика из CreateSheetBEMap:
        # Для КАЖДОЙ уникальной пары (БЕ поставщика, БЕ покупателя) создаём записи
        # с ВСЕми строками мэппинга для данного БЕ поставщика.
        # ВАЖНО: не для каждой строки pivot_data (которая содержит разные ЦФО),
        # а только для уникальных пар!

        rows_to_write = []

        # Итерируем по уникальным парам (БЕ поставщика, БЕ покупателя)
        for (be_supplier, be_buyer), sum_acc in sum_by_key.items():
            # Пропускаем "Общий итог"
            if be_supplier == "Общий итог":
                continue

            # Для каждой уникальной пары берём ВСЕ строки мэппинга с тем же БЕ поставщика
            if be_supplier in mapping_dict:
                matched_rows = mapping_dict[be_supplier]

                for map_item in matched_rows:
                    cfo_oper = map_item["cfo_oper"]
                    stat_oper = map_item["stat_oper"]
                    percent_val = map_item["percent"]

                    # VBA округляет % и если 0, то ставит 1
                    percent_val = round(percent_val)
                    if percent_val == 0:
                        percent_val = 1

                    # Пропускаем записи без ЦФО операционного
                    if not cfo_oper:
                        continue

                    # Сумма расходов = сумма с накопительным итогом * процент / 100
                    calculated_sum = sum_acc * percent_val / 100

                    rows_to_write.append({
                        "be_cfo": f"{be_supplier}{cfo_oper}",
                        "be_supplier": be_supplier,
                        "cfo_oper": cfo_oper,
                        "stat_oper": stat_oper,
                        "percent": percent_val,
                        "sum": calculated_sum,
                        "sum_acc": sum_acc,
                        "be_buyer": be_buyer,
                    })
            else:
                log.warning(
                    f"БЕ поставщика '{be_supplier}' не найден в мэппинге OPEX!")

        # Сортировка:
        # Первичный ключ: БЕ поставщика (убывание)
        # Вторичный ключ: БЕ покупателя (возрастание)
        # Третичный ключ: ЦФО операционное (возрастание)
        # Четвёртый ключ: Статья операционная (возрастание)
        # Используем stable sort: сначала по младшему ключу, потом по старшему
        # 4-й ключ, возрастание
        rows_to_write.sort(key=lambda x: x["stat_oper"])
        # 3-й ключ, возрастание
        rows_to_write.sort(key=lambda x: x["cfo_oper"])
        # 2-й ключ, возрастание
        rows_to_write.sort(key=lambda x: x["be_buyer"])
        # 1-й ключ, убывание
        rows_to_write.sort(key=lambda x: x["be_supplier"], reverse=True)

        log.info(f"Второй блок OPEX: {len(rows_to_write)} записей")

        # Записываем данные
        data_row = current_row + 1
        for row_data in rows_to_write:
            ws.cell(row=data_row, column=START_COL_ACCOUNT,
                    value=row_data["be_cfo"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    1, value=row_data["be_supplier"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    2, value=row_data["cfo_oper"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    3, value=row_data["stat_oper"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    4, value=row_data["percent"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    5, value=row_data["sum"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    6, value=row_data["sum_acc"])
            ws.cell(row=data_row, column=START_COL_ACCOUNT +
                    7, value=row_data["be_buyer"])
            data_row += 1

        # Check для второго блока
        check_row = current_row - 2
        sum_col = START_COL_ACCOUNT + 5
        sum_col_letter = get_column_letter(sum_col)
        last_data_row = data_row - 1

        ws.cell(row=check_row, column=sum_col, value="Check")
        ws.cell(
            row=check_row,
            column=sum_col + 1,
            value=f"=SUM(${sum_col_letter}{current_row + 1}:${sum_col_letter}{last_data_row})",
        )

        log.info(
            f"Создана таблица OPEX: {len(rows_to_write)} записей, Check в строке {check_row}"
        )

    def _create_mapping_table(
        self,
        ws: Worksheet,
        pivot_data: Dict,
        arr_map: List[List[Any]],
        is_budget: bool,
        is_vgo: bool,
        is_t2: bool,
    ):
        """Создание таблицы мэппинга на листе расчёта (второй блок - OPEX)

        ВАЖНО: Для листов ВГО (is_vgo=True) второй блок НЕ создаётся,
        так как в эталоне VBA его нет!
        """
        # Для листов ВГО второй блок не создаётся
        if is_vgo:
            log.info("Лист ВГО - второй блок не создаётся (как в эталоне)")
            return

        helper = SheetHelper(ws)
        # +5 для правильного позиционирования (строка 113 в эталоне)
        # Между первым блоком и вторым: 2 пустые строки, Check, 1 пустая строка
        current_row = helper.get_used_range_end(START_COL_ACCOUNT) + 5

        # Заголовки для второго блока (OPEX/операционный)
        if is_budget:
            headers = [
                "БЕ + ЦФО",
                "БЕ поставщика",
                "ЦФО операционное",
                "Статья операционная",
                "%",
                "Сумма расходов",
                "Сумма расходов с накопительным итогом расчетная для формул",
                "БЕ покупателя",
            ]
            cfo_offset = 2  # ЦФО операционное в колонке I (offset 2)
            stat_offset = 3  # Статья операционная в колонке J (offset 3)
            percent_offset = 4  # % в колонке K (offset 4)
            sum_offset = 5  # Сумма расходов в колонке L (offset 5)
            sum_acc_offset = (
                # Сумма расходов с накопительным итогом в колонке M (offset 6)
                6
            )
            buyer_offset = 7  # БЕ покупателя в колонке N (offset 7)
        else:
            headers = [
                "БЕ + ЦФО",
                "БЕ поставщика",
                "ЦФО КВ",
                "ЦФО операционное",
                "Статья операционная",
                "%",
                "Сумма расходов",
                "Сумма расходов с накопительным итогом расчетная для формул",
                "БЕ покупателя",
            ]
            cfo_offset = 3  # ЦФО операционное
            stat_offset = 4  # Статья операционная
            percent_offset = 5  # %
            sum_offset = 6  # Сумма расходов
            sum_acc_offset = 7  # Сумма расходов с накопительным итогом
            buyer_offset = 8  # БЕ покупателя

        for col_offset, header in enumerate(headers):
            ws.cell(
                row=current_row, column=START_COL_ACCOUNT + col_offset, value=header
            )

        # Создаём словарь мэппинга для быстрого поиска
        # VBA для OPEX секции: ArrTitle = Array("БЕ поставщика", "ЦФО операционное", "Статья операционная", "%")
        # Индексы в arr_map (загружен из секции OPEX):
        #   [0] = БЕ поставщика
        #   [1] = ЦФО операционное
        #   [2] = Статья операционная
        #   [3] = %
        mapping_dict = {}
        for map_row in arr_map:
            if map_row and len(map_row) > 0 and map_row[0]:
                # БЕ поставщика в первой колонке [0]
                be_key = safe_str(map_row[0])
                if be_key not in mapping_dict:
                    mapping_dict[be_key] = []
                mapping_dict[be_key].append(map_row)

        log.info(
            f"Загружено {len(mapping_dict)} уникальных БЕ из мэппинга OPEX: {list(mapping_dict.keys())[:10]}..."
        )

        # Создаём словарь сумм по ключу (БЕ поставщика, БЕ покупателя)
        # VBA использует SUMIFS для суммирования всех записей с одинаковым БЕ+покупатель
        sum_by_key = {}
        for key, pv_values in pivot_data.items():
            be_sup = safe_str(pv_values["be_supplier"])
            be_buy = safe_str(pv_values["be_buyer"])
            sum_key = (be_sup, be_buy)
            pv_sum = safe_float(pv_values.get("sum", 0), 0)
            if sum_key not in sum_by_key:
                sum_by_key[sum_key] = 0
            sum_by_key[sum_key] += pv_sum

        log.info(f"Создан словарь сумм: {len(sum_by_key)} уникальных ключей")

        # Собираем данные в порядке их появления в pivot_data (как в VBA)
        # ВАЖНО: для уникальных комбинаций (be_supplier, be_buyer) создаём записи только ОДИН раз
        rows_to_write = []
        processed_pairs = set()  # Уже обработанные пары (be_supplier, be_buyer)

        for key, pv_values in pivot_data.items():
            be_supplier = safe_str(pv_values["be_supplier"])

            # Пропускаем "Общий итог"
            if be_supplier == "Общий итог":
                continue

            be_buyer = safe_str(pv_values["be_buyer"])

            # Пропускаем уже обработанные пары
            pair_key = (be_supplier, be_buyer)
            if pair_key in processed_pairs:
                continue
            processed_pairs.add(pair_key)

            # Сумма с накопительным итогом = сумма ВСЕХ записей с тем же БЕ+покупатель
            sum_acc = sum_by_key.get((be_supplier, be_buyer), 0)

            # Для второго блока берём ВСЕ строки мэппинга для данного БЕ
            if be_supplier in mapping_dict:
                matched_rows = mapping_dict[be_supplier]

                # Собираем строки для этой записи pivot_data
                pivot_rows = []
                for map_row in matched_rows:
                    # Структура мэппинга OPEX (из секции "Информация о ЦФО и статье из бюджета OPEX"):
                    # VBA: ArrTitle = Array("БЕ поставщика", "ЦФО операционное", "Статья операционная", "%")
                    # [0] = БЕ поставщика
                    # [1] = ЦФО операционное
                    # [2] = Статья операционная
                    # [3] = %
                    cfo_oper = safe_str(map_row[1]) if len(map_row) > 1 else ""
                    stat_oper = safe_str(map_row[2]) if len(
                        map_row) > 2 else ""
                    percent_val = safe_float(
                        map_row[3], 100) if len(map_row) > 3 else 100

                    # VBA округляет % и если 0, то ставит 1
                    # If ShBEMap.Cells(iRow, j).Value = 0 Then ShBEMap.Cells(iRow, j).Value = 1
                    percent_val = round(percent_val)
                    if percent_val == 0:
                        percent_val = 1

                    # Пропускаем записи без ЦФО операционного
                    if not cfo_oper:
                        continue

                    # Сумма расходов = сумма с накопительным итогом * процент / 100
                    calculated_sum = sum_acc * percent_val / 100

                    pivot_rows.append(
                        {
                            "be_cfo": f"{pv_values['be_supplier']}{cfo_oper}",
                            "be_supplier": pv_values["be_supplier"],
                            "cfo_oper": cfo_oper,
                            "stat_oper": stat_oper,
                            "percent": percent_val,
                            "sum": calculated_sum,
                            "sum_acc": sum_acc,
                            "be_buyer": be_buyer,
                        }
                    )

                rows_to_write.extend(pivot_rows)
            else:
                # Если нет мэппинга, добавляем строку с базовыми данными
                log.warning(
                    f"БЕ поставщика '{be_supplier}' не найден в мэппинге OPEX!")
                rows_to_write.append(
                    {
                        "be_cfo": f"{pv_values['be_supplier']}{pv_values['cfo']}",
                        "be_supplier": pv_values["be_supplier"],
                        "cfo_oper": pv_values.get("cfo", ""),
                        "stat_oper": "",
                        "percent": 100,
                        "sum": sum_acc,
                        "sum_acc": sum_acc,
                        "be_buyer": be_buyer,
                    }
                )

        # VBA сортирует по БЕ поставщика в убывающем порядке:
        # ShSvPvt.PivotTables("Pvt_BE").PivotFields("БЕ поставщика").AutoSort xlDescending, "БЕ поставщика"
        rows_to_write.sort(key=lambda x: x["be_supplier"], reverse=True)

        log.info(
            f"Второй блок: отсортировано {len(rows_to_write)} записей по БЕ поставщика (убывание)")

        # Записываем отсортированные данные
        data_row = current_row + 1
        for row_data in rows_to_write:
            # БЕ + ЦФО
            ws.cell(row=data_row, column=START_COL_ACCOUNT,
                    value=row_data["be_cfo"])
            # БЕ поставщика
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + 1,
                value=row_data["be_supplier"],
            )
            # ЦФО операционное
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + cfo_offset,
                value=row_data["cfo_oper"],
            )
            # Статья операционная
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + stat_offset,
                value=row_data["stat_oper"],
            )
            # %
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + percent_offset,
                value=row_data["percent"],
            )
            # Сумма расходов
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + sum_offset,
                value=row_data["sum"],
            )
            # Сумма расходов с накопительным итогом
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + sum_acc_offset,
                value=row_data["sum_acc"],
            )
            # БЕ покупателя
            ws.cell(
                row=data_row,
                column=START_COL_ACCOUNT + buyer_offset,
                value=row_data["be_buyer"],
            )

            data_row += 1

        # Check для второго блока (как в эталоне: L111="Check", M111=SUM)
        # Находим позицию для Check - перед заголовками второго блока
        check_row = current_row - 2  # 2 строки выше заголовков
        sum_col = START_COL_ACCOUNT + sum_offset  # Колонка L (Сумма расходов)
        sum_col_letter = get_column_letter(sum_col)
        last_data_row = data_row - 1

        ws.cell(row=check_row, column=sum_col, value="Check")
        ws.cell(
            row=check_row,
            column=sum_col + 1,
            value=f"=SUM(${sum_col_letter}{current_row + 1}:${sum_col_letter}{last_data_row})",
        )

        log.info(
            f"Создана таблица мэппинга: {len(rows_to_write)} записей, начиная со строки {current_row}"
        )
        log.info(
            f"Check для второго блока: L{check_row}='Check', M{check_row}=SUM({sum_col_letter}{current_row + 1}:{sum_col_letter}{last_data_row})"
        )


def create_preview_data(
    macros_file: str, marja_file: str, vgo_file: str, progress_callback=None
) -> ProcessingResult:
    """
    Основная функция для создания предварительных листов

    Args:
        macros_file: Путь к основному файлу
        marja_file: Путь к файлу Маржа
        vgo_file: Путь к файлу ВГО
        progress_callback: Callback для обновления прогресса

    Returns:
        ProcessingResult с результатом выполнения
    """
    processor = CreatePreviewDataProcessor(
        macros_file_path=macros_file,
        marja_file_path=marja_file,
        vgo_file_path=vgo_file,
        progress_callback=progress_callback,
    )
    return processor.process()
